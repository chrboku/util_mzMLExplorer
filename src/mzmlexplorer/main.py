print("Loading mzmlexplorer...")

import sys
import os
from PyQt6.QtWidgets import (
    QApplication,
    QMainWindow,
    QVBoxLayout,
    QHBoxLayout,
    QWidget,
    QPushButton,
    QTableWidget,
    QTableWidgetItem,
    QTreeWidget,
    QTreeWidgetItem,
    QTreeWidgetItemIterator,
    QLabel,
    QDoubleSpinBox,
    QLineEdit,
    QGroupBox,
    QCheckBox,
    QComboBox,
    QSpinBox,
    QSplitter,
    QFileDialog,
    QMessageBox,
    QHeaderView,
    QMenuBar,
    QMenu,
    QDialog,
    QFormLayout,
    QProgressDialog,
    QFrame,
    QScrollArea,
    QAbstractItemView,
    QWidgetAction,
    QTabWidget,
)
from PyQt6.QtCore import Qt, QTimer, QSettings, QEvent
from PyQt6.QtGui import (
    QFont,
    QAction,
    QDragEnterEvent,
    QDropEvent,
    QCursor,
    QPixmap,
    QPainter,
    QPen,
)
from PyQt6.QtCore import Qt
from PyQt6.QtGui import QColor, QFont
import pandas as pd
from .compound_manager import CompoundManager
from .file_manager import FileManager
from .windows import EICWindow, MultiAdductWindow
from .window_shared import CollapsibleBox, NoScrollComboBox, NoScrollSpinBox, NoScrollDoubleSpinBox
from .window_file_explorer import MzMLFileExplorerWindow
from .window_msms import USISpectrumComparisonWindow
from .compound_import_dialog import (
    CompoundImportDialog,
    validate_formula_smiles_agreement,
)

import json
import toml
import concurrent.futures
import re

# fmt: off
# Full adduct library used both for the compounds template and the custom EIC dialog.
ADDUCTS_TEMPLATE_DATA = [
    ## Pos mode
    {"Adduct": "[M+H]+",            "ElementsAdded": "H"     , "ElementsLost": ""    , "Mass_change": None        , "Charge":  1, "Multiplier": 1},
    {"Adduct": "[M+NH4]+",          "ElementsAdded": "NH4"   , "ElementsLost": ""    , "Mass_change": None        , "Charge":  1, "Multiplier": 1},
    {"Adduct": "[M+Na]+",           "ElementsAdded": "Na"    , "ElementsLost": ""    , "Mass_change": None        , "Charge":  1, "Multiplier": 1},
    {"Adduct": "[M+IsoProp+H]+",    "ElementsAdded": ""      , "ElementsLost": ""    , "Mass_change": 61.0653400  , "Charge":  1, "Multiplier": 1},
    {"Adduct": "[M+CH3OH+H]+",      "ElementsAdded": "CH3OHH", "ElementsLost": ""    , "Mass_change": None        , "Charge":  1, "Multiplier": 1},
    {"Adduct": "[M+K]+",            "ElementsAdded": "K"     , "ElementsLost": ""    , "Mass_change": None        , "Charge":  1, "Multiplier": 1},
    {"Adduct": "[M+ACN+H]+",        "ElementsAdded": ""      , "ElementsLost": ""    , "Mass_change": 42.0338230  , "Charge":  1, "Multiplier": 1},
    {"Adduct": "[M+2Na-H]+",        "ElementsAdded": "Na2"   , "ElementsLost": "H"   , "Mass_change": None        , "Charge":  1, "Multiplier": 1},
    {"Adduct": "[M+ACN+Na]+",       "ElementsAdded": ""      , "ElementsLost": ""    , "Mass_change": 64.0157650  , "Charge":  1, "Multiplier": 1},
    {"Adduct": "[M+2K-H]+",         "ElementsAdded": "K2"    , "ElementsLost": "H"   , "Mass_change": None        , "Charge":  1, "Multiplier": 1},
    {"Adduct": "[M+DMSO+H]+",       "ElementsAdded": ""      , "ElementsLost": ""    , "Mass_change": 79.0212200  , "Charge":  1, "Multiplier": 1},
    {"Adduct": "[M+2ACN+H]+",       "ElementsAdded": ""      , "ElementsLost": ""    , "Mass_change": 83.0603700  , "Charge":  1, "Multiplier": 1},
    {"Adduct": "[M+IsoProp+Na+H]+", "ElementsAdded": ""      , "ElementsLost": ""    , "Mass_change": 84.0551100  , "Charge":  1, "Multiplier": 1},
    {"Adduct": "[2M+H]+",           "ElementsAdded": "H"     , "ElementsLost": ""    , "Mass_change": None        , "Charge":  1, "Multiplier": 2},
    {"Adduct": "[2M+NH4]+",         "ElementsAdded": "NH4"   , "ElementsLost": ""    , "Mass_change": None        , "Charge":  1, "Multiplier": 2},
    {"Adduct": "[2M+Na]+",          "ElementsAdded": "Na"    , "ElementsLost": ""    , "Mass_change": None        , "Charge":  1, "Multiplier": 2},
    {"Adduct": "[2M+K]+",           "ElementsAdded": "K"     , "ElementsLost": ""    , "Mass_change": None        , "Charge":  1, "Multiplier": 2},
    {"Adduct": "[2M+ACN+H]+",       "ElementsAdded": ""      , "ElementsLost": ""    , "Mass_change": 42.0338230  , "Charge":  1, "Multiplier": 2},
    {"Adduct": "[2M+ACN+Na]+",      "ElementsAdded": ""      , "ElementsLost": ""    , "Mass_change": 64.0157650  , "Charge":  1, "Multiplier": 2},
    {"Adduct": "[M+2H]++",          "ElementsAdded": "H2"    , "ElementsLost": ""    , "Mass_change": None        , "Charge":  2, "Multiplier": 1},
    {"Adduct": "[M+H+NH4]++",       "ElementsAdded": "NH4H"  , "ElementsLost": ""    , "Mass_change": None        , "Charge":  2, "Multiplier": 1},
    {"Adduct": "[M+H+Na]++",        "ElementsAdded": "NaH"   , "ElementsLost": ""    , "Mass_change": None        , "Charge":  2, "Multiplier": 1},
    {"Adduct": "[M+H+K]++",         "ElementsAdded": "KH"    , "ElementsLost": ""    , "Mass_change": None        , "Charge":  2, "Multiplier": 1},
    {"Adduct": "[M+ACN+2H]++",      "ElementsAdded": ""      , "ElementsLost": ""    , "Mass_change": 21.5205500  , "Charge":  2, "Multiplier": 1},
    {"Adduct": "[M+2Na]++",         "ElementsAdded": "Na2"   , "ElementsLost": ""    , "Mass_change": None        , "Charge":  2, "Multiplier": 1},
    {"Adduct": "[M+2ACN+2H]++",     "ElementsAdded": ""      , "ElementsLost": ""    , "Mass_change": 42.0338230  , "Charge":  2, "Multiplier": 1},
    {"Adduct": "[M+3ACN+2H]++",     "ElementsAdded": ""      , "ElementsLost": ""    , "Mass_change": 62.5470970  , "Charge":  2, "Multiplier": 1},
    {"Adduct": "[M+3H]+++",         "ElementsAdded": "H3"    , "ElementsLost": ""    , "Mass_change": None        , "Charge":  3, "Multiplier": 1},
    {"Adduct": "[M+2H+Na]+++",      "ElementsAdded": "NaH2"  , "ElementsLost": ""    , "Mass_change": None        , "Charge":  3, "Multiplier": 1},
    {"Adduct": "[M+H+2Na]+++",      "ElementsAdded": "Na2H"  , "ElementsLost": ""    , "Mass_change": None        , "Charge":  3, "Multiplier": 1},
    {"Adduct": "[M+3Na]+++",        "ElementsAdded": "Na3"   , "ElementsLost": ""    , "Mass_change": None        , "Charge":  3, "Multiplier": 1},
    ## Neg mode
    {"Adduct": "[M+Na-2H]-",        "ElementsAdded": "Na"    , "ElementsLost": "H2"  , "Mass_change": None        , "Charge": -1, "Multiplier": 1},
    {"Adduct": "[M+K-2H]-",         "ElementsAdded": "K"     , "ElementsLost": "H2"  , "Mass_change": None        , "Charge": -1, "Multiplier": 1},
    {"Adduct": "[M+Hac-H]-",        "ElementsAdded": ""      , "ElementsLost": ""    , "Mass_change": 59.0138510  , "Charge": -1, "Multiplier": 1},
    {"Adduct": "[M+Br]-",           "ElementsAdded": "Br"    , "ElementsLost": ""    , "Mass_change": None        , "Charge": -1, "Multiplier": 1},
    {"Adduct": "[M+TFA-H]-",        "ElementsAdded": ""      , "ElementsLost": ""    , "Mass_change": 112.985586  , "Charge": -1, "Multiplier": 1},
    {"Adduct": "[M-H]-",            "ElementsAdded": ""      , "ElementsLost": "H"   , "Mass_change": None        , "Charge": -1, "Multiplier": 1},
    {"Adduct": "[M+Cl]-",           "ElementsAdded": "Cl"    , "ElementsLost": ""    , "Mass_change": None        , "Charge": -1, "Multiplier": 1},
    {"Adduct": "[M+FA-H]-",         "ElementsAdded": ""      , "ElementsLost": ""    , "Mass_change": 44.9982010  , "Charge": -1, "Multiplier": 1},
    {"Adduct": "[M-H2O-H]-",        "ElementsAdded": ""      , "ElementsLost": "H2OH", "Mass_change": None        , "Charge": -1, "Multiplier": 1},
    {"Adduct": "[2M-H]-",           "ElementsAdded": ""      , "ElementsLost": "H"   , "Mass_change": None        , "Charge": -1, "Multiplier": 2},
    {"Adduct": "[2M+FA-H]-",        "ElementsAdded": ""      , "ElementsLost": ""    , "Mass_change": 44.9982010  , "Charge": -1, "Multiplier": 2},
    {"Adduct": "[2M+Hac-H]-",       "ElementsAdded": ""      , "ElementsLost": ""    , "Mass_change": 59.0138510  , "Charge": -1, "Multiplier": 2},
    {"Adduct": "[3M-H]-",           "ElementsAdded": ""      , "ElementsLost": "H"   , "Mass_change": None        , "Charge": -1, "Multiplier": 3},
    {"Adduct": "[M-2H]--",          "ElementsAdded": ""      , "ElementsLost": "H2"  , "Mass_change": None        , "Charge": -2, "Multiplier": 1},
    {"Adduct": "[M-3H]---",         "ElementsAdded": ""      , "ElementsLost": "H3"  , "Mass_change": None        , "Charge": -3, "Multiplier": 1},
]
# fmt: on


class CompoundStructurePopup(QWidget):
    """Borderless, non-blocking popup that displays a compound's 2-D structure.

    Shown when the mouse enters a compound row in the substance table and
    hidden when the mouse leaves the table viewport.
    """

    _IMG_SIZE = 240  # structure image side length in pixels

    def __init__(self, parent=None):
        super().__init__(
            parent,
            Qt.WindowType.FramelessWindowHint | Qt.WindowType.Tool | Qt.WindowType.WindowDoesNotAcceptFocus,
        )
        self.setAttribute(Qt.WidgetAttribute.WA_ShowWithoutActivating)

        # Pure white background
        self.setAutoFillBackground(True)
        pal = self.palette()
        pal.setColor(self.backgroundRole(), Qt.GlobalColor.white)
        self.setPalette(pal)

        self.setFixedSize(self._IMG_SIZE + 8, self._IMG_SIZE + 30)

        layout = QVBoxLayout(self)
        layout.setContentsMargins(4, 4, 4, 4)
        layout.setSpacing(2)

        self._formula_label = QLabel()
        self._formula_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self._formula_label.setStyleSheet("font-weight: bold; font-size: 11px;")
        layout.addWidget(self._formula_label)

        self._image_label = QLabel()
        self._image_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self._image_label.setFixedSize(self._IMG_SIZE, self._IMG_SIZE)
        layout.addWidget(self._image_label)

    def paintEvent(self, event):
        """Draw a visible grey border around the popup."""
        super().paintEvent(event)
        painter = QPainter(self)
        pen = QPen(QColor("#888888"))
        pen.setWidth(2)
        painter.setPen(pen)
        painter.drawRect(1, 1, self.width() - 2, self.height() - 2)

    def show_for_compound(self, compound_data: dict, global_pos) -> None:
        """Render *compound_data*'s structure and show the popup near *global_pos*.

        Hides silently when rdkit is unavailable or SMILES is missing/invalid.
        """
        smiles = compound_data.get("SMILES") or compound_data.get("smiles")
        if not smiles:
            self.hide()
            return
        smiles = str(smiles).strip()
        if not smiles or smiles.lower() in ("nan", "none"):
            self.hide()
            return

        # Title text: prefer chemical formula, fall back to compound name
        formula = str(compound_data.get("ChemicalFormula") or compound_data.get("molecular_formula") or "").strip()
        if not formula or formula.lower() in ("nan", "none"):
            formula = str(compound_data.get("Name", "")).strip()
        self._formula_label.setText(formula)

        try:
            from rdkit import Chem
            from rdkit.Chem.Draw import rdMolDraw2D

            mol = Chem.MolFromSmiles(smiles)
            if mol is None:
                self.hide()
                return
            drawer = rdMolDraw2D.MolDraw2DCairo(self._IMG_SIZE, self._IMG_SIZE)
            drawer.drawOptions().clearBackground = True
            drawer.DrawMolecule(mol)
            drawer.FinishDrawing()
            png_bytes = drawer.GetDrawingText()
            pixmap = QPixmap()
            pixmap.loadFromData(png_bytes, "PNG")
            self._image_label.setPixmap(pixmap)
        except Exception:
            self.hide()
            return

        # Position to the right of the cursor; clamp to available screen area
        screen = QApplication.screenAt(global_pos)
        if screen is None:
            screen = QApplication.primaryScreen()
        rect = screen.availableGeometry()
        x = global_pos.x() + 20
        y = global_pos.y() - self.height() // 2
        if x + self.width() > rect.right():
            x = global_pos.x() - self.width() - 10
        y = max(rect.top(), min(y, rect.bottom() - self.height()))
        self.move(x, y)
        self.show()


# Path to the bundled logo image
_LOGO_PATH = os.path.join(os.path.dirname(__file__), "logo.png")


class CustomEICDialog(QDialog):
    """Dialog for quickly plotting an EIC trace for a custom formula, mass, SMILES or m/z value."""

    def __init__(self, adducts_data, ppm_default=5.0, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Show Custom EIC Trace")
        self.setMinimumWidth(500)
        self._adducts_data = adducts_data
        self._ppm_default = ppm_default
        self._result_compound_data = None
        self._result_adduct = None
        self._result_mz = None
        self._result_polarity = None
        self._result_ppm = ppm_default
        self._setup_ui()
        self._connect_signals()
        # Trigger initial validation
        self._validate_tab1()
        self._validate_tab2()

    # ------------------------------------------------------------------
    # UI construction
    # ------------------------------------------------------------------

    def _setup_ui(self):
        layout = QVBoxLayout(self)

        self.tabs = QTabWidget()

        # ---- Tab 0 (default): m/z value ----
        tab2 = QWidget()
        form2 = QFormLayout(tab2)
        form2.setLabelAlignment(Qt.AlignmentFlag.AlignRight)
        form2.setContentsMargins(12, 12, 12, 12)
        form2.setVerticalSpacing(8)

        self.mz_spin = NoScrollDoubleSpinBox()
        self.mz_spin.setRange(0.001, 100000.0)
        self.mz_spin.setDecimals(6)
        self.mz_spin.setSuffix(" m/z")
        self.mz_spin.setValue(200.0)
        form2.addRow("m/z value:", self.mz_spin)

        self.polarity_combo = NoScrollComboBox()
        self.polarity_combo.addItem("Positive", "positive")
        self.polarity_combo.addItem("Negative", "negative")
        form2.addRow("Polarity:", self.polarity_combo)

        self.tabs.addTab(tab2, "m/z value")

        # ---- Tab 1: Formula / SMILES / Mass ----
        tab1 = QWidget()
        form1 = QFormLayout(tab1)
        form1.setLabelAlignment(Qt.AlignmentFlag.AlignRight)
        form1.setContentsMargins(12, 12, 12, 12)
        form1.setVerticalSpacing(8)

        self.formula_edit = QLineEdit()
        self.formula_edit.setPlaceholderText("e.g. C6H12O6")
        form1.addRow("Chemical Formula:", self.formula_edit)

        self.smiles_edit = QLineEdit()
        self.smiles_edit.setPlaceholderText("Optional \u2013 validated with RDKit when filled")
        form1.addRow("SMILES:", self.smiles_edit)

        self.mass_spin = NoScrollDoubleSpinBox()
        self.mass_spin.setRange(0.0, 100000.0)
        self.mass_spin.setDecimals(6)
        self.mass_spin.setSuffix(" Da")
        self.mass_spin.setValue(0.0)
        self.mass_spin.setSpecialValueText("(derive from formula)")
        self.mass_spin.setToolTip("Enter a monoisotopic neutral mass directly. Leave at 0 to derive it automatically from the chemical formula above.")
        form1.addRow("Neutral Mass:", self.mass_spin)

        self.adduct_combo1 = NoScrollComboBox()
        self._fill_adducts_combo(self.adduct_combo1)
        form1.addRow("Adduct:", self.adduct_combo1)

        self.mz_preview_label = QLabel("(enter formula or mass + adduct to preview m/z)")
        self.mz_preview_label.setStyleSheet("color: gray; font-style: italic;")
        form1.addRow("Calculated m/z:", self.mz_preview_label)

        self.polarity_preview_label = QLabel("")
        self.polarity_preview_label.setStyleSheet("color: gray; font-style: italic;")
        form1.addRow("Polarity:", self.polarity_preview_label)

        self.validation_label1 = QLabel("")
        self.validation_label1.setWordWrap(True)
        form1.addRow("", self.validation_label1)

        self.tabs.addTab(tab1, "Formula / SMILES / Mass")

        layout.addWidget(self.tabs)

        # ---- Shared ppm tolerance spinner ----
        ppm_layout = QHBoxLayout()
        ppm_label = QLabel("m/z Tolerance:")
        self.ppm_spin = NoScrollDoubleSpinBox()
        self.ppm_spin.setRange(0.1, 500.0)
        self.ppm_spin.setDecimals(1)
        self.ppm_spin.setSingleStep(1.0)
        self.ppm_spin.setSuffix(" ppm")
        self.ppm_spin.setValue(self._ppm_default)
        self.ppm_spin.setToolTip("m/z extraction window in parts-per-million")
        ppm_layout.addWidget(ppm_label)
        ppm_layout.addWidget(self.ppm_spin)
        ppm_layout.addStretch()
        layout.addLayout(ppm_layout)

        # Buttons
        btn_layout = QHBoxLayout()
        btn_layout.addStretch()
        self.ok_btn = QPushButton("Show EIC")
        self.ok_btn.setEnabled(False)
        self.ok_btn.setDefault(True)
        cancel_btn = QPushButton("Cancel")
        btn_layout.addWidget(self.ok_btn)
        btn_layout.addWidget(cancel_btn)
        layout.addLayout(btn_layout)

        self.ok_btn.clicked.connect(self._on_accept)
        cancel_btn.clicked.connect(self.reject)

    def _fill_adducts_combo(self, combo):
        """Populate an adduct combo-box from the adducts table."""
        if self._adducts_data is not None and not self._adducts_data.empty:
            adducts = self._adducts_data["Adduct"].tolist()
        else:
            adducts = [row["Adduct"] for row in ADDUCTS_TEMPLATE_DATA]
        for adduct in adducts:
            combo.addItem(adduct)

    # ------------------------------------------------------------------
    # Signal connections / validation
    # ------------------------------------------------------------------

    def _connect_signals(self):
        self.formula_edit.textChanged.connect(self._validate_tab1)
        self.smiles_edit.textChanged.connect(self._validate_tab1)
        self.mass_spin.valueChanged.connect(self._validate_tab1)
        self.adduct_combo1.currentIndexChanged.connect(self._validate_tab1)
        self.mz_spin.valueChanged.connect(self._validate_tab2)
        self.polarity_combo.currentIndexChanged.connect(self._validate_tab2)
        self.tabs.currentChanged.connect(self._on_tab_changed)

    def _get_adduct_polarity(self, adduct: str):
        """Return 'positive', 'negative', or None for the given adduct string."""
        # Deduce from trailing character first (most reliable)
        stripped = adduct.rstrip()
        if stripped.endswith("+"):
            return "positive"
        if stripped.endswith("-"):
            return "negative"
        # Fall back to adducts table
        if self._adducts_data is not None and not self._adducts_data.empty:
            row = self._adducts_data[self._adducts_data["Adduct"] == adduct]
            if not row.empty:
                try:
                    charge = int(row.iloc[0]["Charge"])
                    return "positive" if charge > 0 else "negative"
                except Exception:
                    pass
        return None

    def _validate_tab1(self):
        """Live-validate the Formula/SMILES/Mass tab and update feedback labels."""
        from .utils import calculate_molecular_mass, parse_molecular_formula, adduct_mass_change

        formula = self.formula_edit.text().strip()
        smiles = self.smiles_edit.text().strip()
        mass_override = self.mass_spin.value()
        adduct = self.adduct_combo1.currentText()

        errors = []
        warnings = []
        neutral_mass = None

        # --- Validate formula ---
        if formula:
            try:
                parse_molecular_formula(formula)
                neutral_mass = calculate_molecular_mass(formula)
                self.formula_edit.setStyleSheet("QLineEdit { border: 1.5px solid #3a3; }")
            except Exception as exc:
                errors.append(f"Invalid formula: {exc}")
                self.formula_edit.setStyleSheet("QLineEdit { border: 1.5px solid red; }")
        else:
            self.formula_edit.setStyleSheet("")

        # --- Validate SMILES ---
        if smiles:
            try:
                from rdkit import Chem

                mol = Chem.MolFromSmiles(smiles)
                if mol is None:
                    errors.append("SMILES is invalid (RDKit could not parse it)")
                    self.smiles_edit.setStyleSheet("QLineEdit { border: 1.5px solid red; }")
                else:
                    self.smiles_edit.setStyleSheet("QLineEdit { border: 1.5px solid #3a3; }")
                    # Cross-check with formula when both are present and formula is valid
                    if formula and neutral_mass is not None:
                        try:
                            from rdkit.Chem import rdMolDescriptors
                            from .FormulaTools import formulaTools

                            rdkit_formula_str = rdMolDescriptors.CalcMolFormula(mol)
                            ft = formulaTools()
                            rdkit_elems = {k: v for k, v in ft.parseFormula(rdkit_formula_str).items() if k[0].isalpha()}
                            formula_elems = {k: v for k, v in ft.parseFormula(formula).items() if k[0].isalpha()}
                            if rdkit_elems != formula_elems:
                                warnings.append(f"Formula/SMILES mismatch: formula implies {formula}, SMILES implies {rdkit_formula_str}")
                        except Exception:
                            pass
            except ImportError:
                warnings.append("RDKit not available – SMILES cannot be validated")
                self.smiles_edit.setStyleSheet("")
        else:
            self.smiles_edit.setStyleSheet("")

        # --- Use mass override when formula is absent or invalid ---
        if neutral_mass is None and mass_override > 0.0:
            neutral_mass = mass_override

        # --- Calculate preview m/z ---
        mz_value = None
        polarity = None
        if neutral_mass is not None and adduct:
            if self._adducts_data is not None and not self._adducts_data.empty:
                adduct_row = self._adducts_data[self._adducts_data["Adduct"] == adduct]
                if not adduct_row.empty:
                    try:
                        mass_change, charge, multiplier = adduct_mass_change(adduct_row.iloc[0])
                        mz_value = (multiplier * neutral_mass + mass_change) / abs(charge)
                        polarity = self._get_adduct_polarity(adduct)
                    except Exception as exc:
                        warnings.append(f"m/z calculation error: {exc}")
            else:
                warnings.append("No adducts table loaded – cannot calculate m/z")

        # --- Update preview labels ---
        if mz_value is not None:
            self.mz_preview_label.setText(f"{mz_value:.6f}")
            self.mz_preview_label.setStyleSheet("color: #1a7a1a; font-weight: bold;")
        else:
            self.mz_preview_label.setText("(enter formula or mass + adduct to preview m/z)")
            self.mz_preview_label.setStyleSheet("color: gray; font-style: italic;")

        if polarity is not None:
            self.polarity_preview_label.setText(polarity.capitalize())
            self.polarity_preview_label.setStyleSheet("color: #1a1a9a; font-weight: bold;" if polarity == "positive" else "color: #9a1a1a; font-weight: bold;")
        else:
            self.polarity_preview_label.setText("")

        # --- Feedback message ---
        if errors:
            self.validation_label1.setText("\u274c " + "; ".join(errors))
            self.validation_label1.setStyleSheet("color: red;")
        elif warnings:
            self.validation_label1.setText("\u26a0\ufe0f " + "; ".join(warnings))
            self.validation_label1.setStyleSheet("color: #b86000;")
        else:
            self.validation_label1.setText("")

        # Enable OK only when on this tab and m/z could be computed without blocking errors
        if self.tabs.currentIndex() == 1:
            self.ok_btn.setEnabled(bool(not errors and mz_value is not None))

    def _validate_tab2(self):
        """Live-validate the m/z tab."""
        valid = self.mz_spin.value() > 0.0
        if self.tabs.currentIndex() == 0:
            self.ok_btn.setEnabled(valid)

    def _on_tab_changed(self, index):
        if index == 0:
            self._validate_tab2()
        else:
            self._validate_tab1()

    # ------------------------------------------------------------------
    # Accept handler
    # ------------------------------------------------------------------

    def _on_accept(self):
        """Collect results and close the dialog."""
        from .utils import calculate_molecular_mass, parse_molecular_formula, adduct_mass_change

        if self.tabs.currentIndex() == 1:
            # ---- Tab 1: Formula / SMILES / Mass ----
            formula = self.formula_edit.text().strip()
            mass_override = self.mass_spin.value()
            adduct = self.adduct_combo1.currentText()

            neutral_mass = None
            if formula:
                try:
                    neutral_mass = calculate_molecular_mass(formula)
                except Exception:
                    pass
            if neutral_mass is None and mass_override > 0.0:
                neutral_mass = mass_override

            adduct_row = self._adducts_data[self._adducts_data["Adduct"] == adduct]
            mass_change, charge, multiplier = adduct_mass_change(adduct_row.iloc[0])
            mz_value = (multiplier * neutral_mass + mass_change) / abs(charge)
            polarity = self._get_adduct_polarity(adduct)

            name = formula if formula else f"Mass {neutral_mass:.4f} Da"
            self._result_compound_data = {
                "Name": name,
                "ChemicalFormula": formula if formula else None,
                "Mass": neutral_mass,
                "RT_min": 50.0,
                "RT_start_min": 0.0,
                "RT_end_min": 100.0,
            }
            self._result_adduct = adduct
            self._result_mz = mz_value
            self._result_polarity = polarity
        else:
            # ---- Tab 0: m/z value ----
            mz_value = self.mz_spin.value()
            polarity = self.polarity_combo.currentData()
            pol_sign = "+" if polarity == "positive" else "-"
            self._result_compound_data = {
                "Name": f"Custom m/z {mz_value:.6f}",
                "RT_min": 50.0,
                "RT_start_min": 0.0,
                "RT_end_min": 100.0,
            }
            self._result_adduct = f"[custom]{pol_sign}"
            self._result_mz = mz_value
            self._result_polarity = polarity

        self._result_ppm = self.ppm_spin.value()
        self.accept()

    def get_result(self):
        """Return (compound_data, adduct, mz_value, polarity, ppm) after the dialog was accepted."""
        return (
            self._result_compound_data,
            self._result_adduct,
            self._result_mz,
            self._result_polarity,
            self._result_ppm,
        )


class MzMLExplorerMainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("mzML Explorer")
        self.setGeometry(100, 100, 1200, 800)

        # Enable drag and drop
        self.setAcceptDrops(True)

        # Initialize settings
        # On Windows, use IniFormat in AppData/Local/mzmlexplorer for a predictable location.
        # On other platforms keep the default native format.
        import sys as _sys
        if _sys.platform == "win32":
            import os as _os
            _appdata = _os.environ.get("LOCALAPPDATA", _os.path.expanduser("~"))
            _ini_path = _os.path.join(_appdata, "mzmlexplorer", "settings.ini")
            _os.makedirs(_os.path.dirname(_ini_path), exist_ok=True)
            self.settings = QSettings(_ini_path, QSettings.Format.IniFormat)
        else:
            self.settings = QSettings("mzMLExplorer", "mzMLExplorer")

        # Data storage (initialize before loading settings that depend on it)
        self.file_manager = FileManager()
        self.compound_manager = CompoundManager()
        self.eic_windows = []

        # Peak integration data storage
        # Key: (compound_name, ion_name) -> dict with integration data
        self.peak_integration_data = {}

        # Compound file monitoring
        self.compound_file_path = None
        self.compound_file_size = None
        self.compound_file_mtime = None
        self.compound_file_monitor_timer = None
        self.compound_file_monitoring_active = False

        # Load settings after data storage is initialized
        self.load_eic_defaults()

        self.init_ui()
        self.load_stylesheet()

    def init_ui(self):
        """Initialize the user interface"""
        # Create menu bar
        self.create_menu_bar()

        central_widget = QWidget()
        self.setCentralWidget(central_widget)

        # Main layout
        main_layout = QVBoxLayout(central_widget)

        # Create splitter for main content
        splitter = QSplitter(Qt.Orientation.Horizontal)

        # Left panel: Files table
        left_panel = QGroupBox("Loaded Files")
        left_panel.setAcceptDrops(True)
        left_layout = QVBoxLayout(left_panel)

        self.files_table = QTableWidget()
        self.files_table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.files_table.setAcceptDrops(True)
        self.files_table.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.files_table.customContextMenuRequested.connect(self.show_files_context_menu)
        self.files_table.verticalHeader().setDefaultSectionSize(20)
        self.files_table.verticalHeader().setMinimumSectionSize(16)
        self.files_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        left_layout.addWidget(self.files_table)

        # Watermark logo shown until files are loaded
        self._logo_overlay = None
        self._setup_logo_overlay()

        # Placeholder hint
        self._files_placeholder = QLabel("Drag 'n' drop files here", self.files_table.viewport())
        self._files_placeholder.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self._files_placeholder.setStyleSheet("color: #aaaaaa; font-size: 13px;")
        self._files_placeholder.setAttribute(Qt.WidgetAttribute.WA_TransparentForMouseEvents)
        self._files_placeholder.show()
        self.files_table.viewport().installEventFilter(self)

        # Right panel: Compounds tree
        right_panel = QGroupBox("Compounds")
        right_panel.setAcceptDrops(True)
        right_layout = QVBoxLayout(right_panel)

        # Custom EIC trace button
        custom_eic_btn = QPushButton("Show Custom EIC Trace")
        custom_eic_btn.setToolTip("Open a dialog to plot an EIC for a custom formula, mass, SMILES or m/z value")
        custom_eic_btn.clicked.connect(self.show_custom_eic_dialog)
        right_layout.addWidget(custom_eic_btn)

        # Add filter line
        filter_layout = QHBoxLayout()
        filter_label = QLabel("Filter:")
        self.compound_filter = QLineEdit()
        self.compound_filter.setPlaceholderText("mz 100-200 | rt 5-10 | SMARTS:[c,C]1([c,C][c,C][c,C][c,C]2)[c,C]2[c,C][c,C][c,C][o,O]1 | name regex")
        self.compound_filter.textChanged.connect(self.filter_compounds)
        filter_layout.addWidget(filter_label)
        filter_layout.addWidget(self.compound_filter)
        right_layout.addLayout(filter_layout)

        self.compounds_table = QTreeWidget()
        self.compounds_table.setColumnCount(7)
        self.compounds_table.setHeaderLabels(["Name", "Retention Time", "Type", "Quantification", "Formula", "Mass", "Common Adducts"])
        self.compounds_table.setSelectionBehavior(QTreeWidget.SelectionBehavior.SelectRows)
        self.compounds_table.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.compounds_table.customContextMenuRequested.connect(self.show_compound_context_menu)
        self.compounds_table.setAcceptDrops(True)
        self.compounds_table.setSortingEnabled(False)  # Disable sorting to maintain group structure

        # Configure tree headers
        header = self.compounds_table.header()
        header.setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        header.setStretchLastSection(False)
        # Set initial column widths
        header.resizeSection(0, 250)
        header.resizeSection(1, 100)
        header.resizeSection(2, 55)
        header.resizeSection(3, 90)
        header.resizeSection(4, 100)
        header.resizeSection(5, 60)
        header.resizeSection(6, 150)

        right_layout.addWidget(self.compounds_table)

        # Placeholder hint
        self._compounds_placeholder = QLabel("Drag 'n' drop compounds table here", self.compounds_table.viewport())
        self._compounds_placeholder.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self._compounds_placeholder.setStyleSheet("color: #aaaaaa; font-size: 13px;")
        self._compounds_placeholder.setAttribute(Qt.WidgetAttribute.WA_TransparentForMouseEvents)
        self._compounds_placeholder.show()
        self.compounds_table.viewport().installEventFilter(self)

        # Structure hover popup — one instance, reused for every hovered row
        self.structure_popup = CompoundStructurePopup(self)
        self.compounds_table.setMouseTracking(True)
        self.compounds_table.viewport().setMouseTracking(True)
        self.compounds_table.itemEntered.connect(self._on_compound_hover)
        self.compounds_table.viewport().installEventFilter(self)

        splitter.addWidget(left_panel)
        splitter.addWidget(right_panel)
        splitter.setSizes([300, 900])

        main_layout.addWidget(splitter)

        # Create memory usage label for bottom left corner
        self.memory_label = QLabel("Memory: -- MB")
        self.memory_label.setStyleSheet("QLabel { font-size: 8px; color: #666; }")

        # Add memory label to status bar
        self.statusBar().addPermanentWidget(self.memory_label)
        self.statusBar().showMessage("Ready")

        # Timer for updating memory usage
        self.memory_timer = QTimer()
        self.memory_timer.timeout.connect(self.update_memory_label)
        self.memory_timer.start(2000)  # Update every 2 seconds

    def load_files(self):
        """Load mzML files from a TSV or Excel file"""
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "Load File List",
            "",
            "Excel files (*.xlsx);;TSV files (*.tsv);;CSV files (*.csv)",
        )

        if file_path:
            self._load_files_from_path(file_path, source="menu")

    def _load_files_from_path(self, file_path, source="menu"):
        """
        Shared method to load files from a file path.

        Args:
            file_path: Path to the file to load
            source: Source of the load ("menu" or "drag & drop")
        """
        try:
            # Load the file list
            if file_path.endswith(".xlsx"):
                df = pd.read_excel(file_path)
            elif file_path.endswith(".tsv"):
                df = pd.read_csv(file_path, sep="\t")
            else:
                df = pd.read_csv(file_path)

            # Process the dataframe (validation + forward-fill + loading)
            self._process_files_dataframe(df, source=source, excel_path=file_path)

        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to load files: {str(e)}")

    def _process_files_dataframe(self, df, source="menu", excel_path=None):
        """
        Shared method to validate and process files dataframe.

        Args:
            df: DataFrame with file list data
            source: Source of the load ("menu" or "drag & drop")

        Returns:
            bool: True if successful, False otherwise
        """
        # Normalize column names (case-insensitive)
        _FILES_COLUMNS = {
            "filepath": "Filepath",
            "group": "group",
            "color": "color",
            "dilution": "Dilution",
            "quantification": "Quantification",
            "sample_id": "sample_id",
            "batch": "batch",
            "injection_order": "injection_order",
            "injection_volume": "injection_volume",
        }
        df = df.rename(columns={col: _FILES_COLUMNS.get(col.lower(), col) for col in df.columns})

        # Validate required columns
        if "Filepath" not in df.columns:
            QMessageBox.warning(self, "Error", "The file must contain a 'Filepath' column!")
            return False

        # Validate that the first row is fully filled
        if df.empty:
            QMessageBox.warning(self, "Error", "The file list is empty!")
            return False

        first_row = df.iloc[0]
        optional_cols = {"sample_id", "Quantification", "Dilution"}
        missing_in_first = first_row.isna()
        required_missing = missing_in_first[~missing_in_first.index.isin(optional_cols)]
        if required_missing.any():
            missing_cols = ", ".join(required_missing[required_missing].index.tolist())
            QMessageBox.warning(
                self,
                "Error",
                f"The first row must be fully filled. Missing values in columns: {missing_cols}",
            )
            return False

        # Validate that Filepath column is fully filled
        if df["Filepath"].isna().any():
            empty_rows = df[df["Filepath"].isna()].index.tolist()
            row_numbers = ", ".join([str(i + 2) for i in empty_rows])  # +2 for 1-based and header
            QMessageBox.warning(
                self,
                "Error",
                f"The 'Filepath' column must be fully filled. Empty cells found in rows: {row_numbers}",
            )
            return False

        # Forward-fill only group and color columns.
        # All other columns (e.g. dilution, injection_volume, quantification, batch)
        # are sample-specific and must not be auto-filled.
        ffill_cols = {"group", "color"}
        for col in df.columns:
            if col in ffill_cols:
                df[col] = df[col].ffill()

        # Load files using file manager
        self.file_manager.load_files(df, excel_path=excel_path)
        self.update_files_table()

        # If memory mode is enabled, load the new files into memory with progress
        if self.file_manager.keep_in_memory:
            self.load_files_to_memory_with_progress()

        total_files = len(self.file_manager.get_files_data())
        source_text = "via drag & drop" if source == "drag & drop" else ""
        self.statusBar().showMessage(f"Files loaded {source_text}. Total: {total_files} files".strip())

        return True

    def load_compounds(self):
        """Load compounds from a file"""
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "Load Compounds",
            "",
            "All supported (*.xlsx *.csv *.tsv);;Excel files (*.xlsx);;CSV files (*.csv);;TSV files (*.tsv)",
        )

        if file_path:
            self.load_compounds_from_file(file_path)

    def clear_compounds(self):
        """Clear all loaded compounds"""
        if self.compound_manager.get_compounds_data().empty:
            QMessageBox.information(self, "Information", "No compounds loaded to clear.")
            return

        reply = QMessageBox.question(
            self,
            "Clear Compounds",
            "Are you sure you want to clear all loaded compounds?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No,
        )

        if reply == QMessageBox.StandardButton.Yes:
            # Stop monitoring compound file
            self.stop_compound_file_monitoring()

            # Clear compounds data
            self.compound_manager.compounds_data = pd.DataFrame()
            self.compound_manager.compound_adduct_data = {}

            # Update the UI
            self.update_compounds_table()

            # Update status
            self.statusBar().showMessage("All compounds cleared.")

    def clear_files(self):
        """Clear all loaded files"""
        if self.file_manager.get_files_data().empty:
            QMessageBox.information(self, "Information", "No files loaded to clear.")
            return

        reply = QMessageBox.question(
            self,
            "Clear Files",
            "Are you sure you want to clear all loaded files?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No,
        )

        if reply == QMessageBox.StandardButton.Yes:
            # Clear files data
            self.file_manager.files_data = pd.DataFrame()

            # Clear any cached file data in memory
            if hasattr(self.file_manager, "files_in_memory"):
                self.file_manager.files_in_memory = {}

            # Update the UI
            self.update_files_table()

            # Update status
            self.statusBar().showMessage("All files cleared.")

    def generate_templates(self):
        """Generate template Excel files for file list and compounds"""
        try:
            # Ask user where to save templates
            save_dir = QFileDialog.getExistingDirectory(self, "Select Directory to Save Templates", "")

            if not save_dir:
                return

            # Generate file list template
            files_template_data = {
                "Filepath": [
                    "C:\\path\\to\\your\\file1.mzML",
                    "C:\\path\\to\\your\\file2.mzML",
                    "C:\\path\\to\\your\\file3.mzML",
                    "C:\\path\\to\\your\\file4.mzML",
                    "C:\\path\\to\\your\\file5.mzML",
                    "C:\\path\\to\\your\\file6.mzML",
                ],
                "group": ["Control", "", "", "Treatment", "", ""],
                "color": ["#1f77b4", "", "", "#ff7f0e", "", ""],
                "batch": ["Batch1", "", "", "Batch1", "", "Batch2"],
                "injection_order": [1, 2, 3, 4, 5, 6],
                "injection_volume": [5.0, "", "", 5.0, "", ""],
                "sample_id": [
                    "CTL_001",
                    "CTL_002",
                    "CTL_003",
                    "TRT_001",
                    "TRT_002",
                    "TRT_003",
                ],
                "quantification": [
                    "{'Caffeine': [500.0, 'ng/mL']}",
                    "{'Glucose': [250.0, 'µM']}",
                    "{'Caffeine': [250.0, 'ng/mL']}",
                    "{'Glucose': [125.0, 'µM']}",
                    "",
                    "",
                ],
            }

            files_template_df = pd.DataFrame(files_template_data)
            files_template_path = os.path.join(save_dir, "file_list_template.xlsx")
            files_template_df.to_excel(files_template_path, index=False)

            # Generate compounds template
            compounds_template_data = {
                "Group": [
                    "Methylxanthines",
                    "Methylxanthines",
                    "",  # No group
                    "Test Group A;Test Group B",  # Multiple groups
                    "",  # No group
                ],
                "Name": [
                    "Caffeine",
                    "Theophylline",
                    "Unknown_Compound_1",
                    "Unknown_Compound_2",
                    "Unknown_No_RT",
                ],
                "ChemicalFormula": [
                    "C8H10N4O2",
                    "C7H8N4O2",
                    "",  # Empty for mass-based compound
                    "",  # Empty for mass-based compound
                    "C10H15N5O",  # Compound without RT info
                ],
                "Mass": [
                    "",  # Empty for formula-based compound
                    "",  # Empty for formula-based compound
                    194.0579,  # Mass-based compound
                    256.1234,  # Mass-based compound
                    "",
                ],
                "RT_min": [5.2, 4.8, 3.1, 7.5, ""],  # Empty RT for last compound
                "RT_start_min": [4.8, 4.4, 2.7, 7.0, ""],  # Empty for last compound
                "RT_end_min": [5.6, 5.2, 3.5, 8.0, ""],  # Empty for last compound
                "Common_adducts": [
                    "[M+H]+, [M+Na]+, [M+K]+",
                    "[M+H]+, [M+Na]+, [M-H]-",
                    "[M+H]+, [195.0652]+",  # Mix of standard and custom m/z
                    "[257.1307]+, [255.1151]-",  # Custom m/z values only
                    "[M+H]+",  # Compound without RT will use full range (0-100 min)
                ],
                "compound_class": [
                    "Alkaloid",
                    "Alkaloid",
                    "Unknown",
                    "Unknown",
                    "Unknown",
                ],
                "cas_number": ["58-08-2", "58-55-9", "", "", ""],
            }

            compounds_template_df = pd.DataFrame(compounds_template_data)

            # Create adducts template
            adducts_template_data = ADDUCTS_TEMPLATE_DATA

            adducts_template_df = pd.DataFrame.from_dict(adducts_template_data)

            # Save to Excel with multiple sheets
            compounds_template_path = os.path.join(save_dir, "compounds_template.xlsx")
            with pd.ExcelWriter(compounds_template_path, engine="openpyxl") as writer:
                compounds_template_df.to_excel(writer, sheet_name="Compounds", index=False)
                adducts_template_df.to_excel(writer, sheet_name="Adducts", index=False)

            # Show success message
            QMessageBox.information(
                self,
                "Templates Generated",
                f"Template files have been generated successfully:\n\n"
                f"• File List Template: {files_template_path}\n"
                f"• Compounds Template: {compounds_template_path}\n\n"
                f"Instructions:\n"
                f"• In the file list template:\n"
                f"  - 'Filepath' column: Full path to your mzML files (required)\n"
                f"  - 'group' column: Group names for organizing samples\n"
                f"  - 'color' column: Hex colors for visualization (optional)\n"
                f"  - Other columns: Additional metadata as needed\n\n"
                f"• In the compounds template:\n"
                f"  - Use either 'ChemicalFormula' OR 'Mass' column\n"
                f"  - 'Common_adducts': Standard adducts (e.g., [M+H]+) or\n"
                f"    custom m/z values (e.g., [197.1234]+, [255.0987]-)\n"
                f"  - 'Adducts' sheet: Defines standard adduct properties\n\n"
                f"• Files will be sorted by group first, then by filename\n"
                f"• The table will show filename first, then metadata, then full path\n\n"
                f"Please edit these files with your data before loading.",
            )

            self.statusBar().showMessage("Template files generated successfully")

        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to generate templates: {str(e)}")

    def update_files_table(self):
        """Update the files table with loaded data"""
        files_display_data = self.file_manager.get_files_display_data()

        if files_display_data.empty:
            self.files_table.setRowCount(0)
            self.files_table.setColumnCount(0)
            return

        # Hide the logo watermark now that files have been loaded
        if self._logo_overlay is not None:
            self._logo_overlay.hide()
            self._logo_overlay.deleteLater()
            self._logo_overlay = None

        # Hide the placeholder once rows exist
        if hasattr(self, "_files_placeholder"):
            self._files_placeholder.setVisible(files_display_data.empty)

        # Set up table
        self.files_table.verticalHeader().setDefaultSectionSize(20)
        self.files_table.setRowCount(len(files_display_data))
        self.files_table.setColumnCount(len(files_display_data.columns))
        self.files_table.setHorizontalHeaderLabels(files_display_data.columns.tolist())

        # Get original files data for group information
        files_data = self.file_manager.get_files_data()

        # Populate table
        for i, (index, row) in enumerate(files_display_data.iterrows()):
            for j, (col_name, value) in enumerate(row.items()):
                display_text = "" if (value is None or (isinstance(value, float) and pd.isna(value))) else str(value)
                item = QTableWidgetItem(display_text)

                # Special handling for color column - only show background color
                if col_name == "color":
                    # Don't show the color text, just the background
                    item.setText("")
                    if pd.notna(value) and value:
                        item.setBackground(QColor(str(value)))
                else:
                    # For other columns, set background color based on group
                    if "group" in row.index and pd.notna(row["group"]):
                        group_color = self.file_manager.get_group_color(row["group"])
                        if group_color and col_name != "color":
                            # Apply a lighter version of the group color for non-color columns
                            color = QColor(group_color)
                            color.setAlpha(50)  # Make it semi-transparent
                            item.setBackground(color)

                self.files_table.setItem(i, j, item)

        # Adjust column widths - allow interactive resizing
        self.files_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        self.files_table.horizontalHeader().setStretchLastSection(True)

    def _get_quant_range_for_compound(self, compound_name: str) -> str:
        """Return a formatted quantification range string for a compound, e.g. '50–200 ng/mL', or '' if none."""
        import json

        files_data = self.file_manager.get_files_data()
        if files_data.empty or "Quantification" not in files_data.columns:
            return ""
        values_by_unit: dict = {}
        for quant_str in files_data["Quantification"].dropna():
            quant_str = str(quant_str).strip()
            if not quant_str:
                continue
            try:
                quant_dict = json.loads(quant_str)
            except (json.JSONDecodeError, ValueError):
                continue
            if compound_name in quant_dict:
                entry = quant_dict[compound_name]
                if isinstance(entry, (list, tuple)) and len(entry) >= 2:
                    try:
                        val = float(entry[0])
                        unit = str(entry[1])
                        values_by_unit.setdefault(unit, []).append(val)
                    except (TypeError, ValueError):
                        pass
                elif isinstance(entry, (int, float)):
                    values_by_unit.setdefault("", []).append(float(entry))
        if not values_by_unit:
            return ""
        parts = []
        for unit, vals in values_by_unit.items():
            lo, hi = min(vals), max(vals)
            if lo == hi:
                parts.append(f"{lo:g} {unit}".strip())
            else:
                parts.append(f"{lo:g}–{hi:g} {unit}".strip())
        return "; ".join(parts)

    # ------------------------------------------------------------------
    # Compound table hover  — structure popup
    # ------------------------------------------------------------------

    def eventFilter(self, obj, event):
        """Hide the structure popup when the mouse leaves the compounds table."""
        if obj is self.compounds_table.viewport() and event.type() == QEvent.Type.Leave:
            self.structure_popup.hide()
        if obj is self.files_table and event.type() == QEvent.Type.Resize:
            self._reposition_logo_overlay()
        # Keep placeholder labels filling their viewport
        if event.type() == QEvent.Type.Resize:
            if obj is self.files_table.viewport() and hasattr(self, "_files_placeholder"):
                self._files_placeholder.setGeometry(obj.rect())
            if obj is self.compounds_table.viewport() and hasattr(self, "_compounds_placeholder"):
                self._compounds_placeholder.setGeometry(obj.rect())
        return super().eventFilter(obj, event)

    def _setup_logo_overlay(self):
        """Pin a semi-transparent logo to the bottom-left of the files table."""
        if not os.path.exists(_LOGO_PATH):
            return
        source_px = QPixmap(_LOGO_PATH)
        scaled = source_px.scaledToWidth(180, Qt.TransformationMode.SmoothTransformation)
        faded = QPixmap(scaled.size())
        faded.fill(Qt.GlobalColor.transparent)
        painter = QPainter(faded)
        painter.setOpacity(0.40)
        painter.drawPixmap(0, 0, scaled)
        painter.end()
        lbl = QLabel(self.files_table)
        lbl.setPixmap(faded)
        lbl.setAttribute(Qt.WidgetAttribute.WA_TransparentForMouseEvents)
        lbl.setFixedSize(faded.size())
        lbl.raise_()
        lbl.show()
        self._logo_overlay = lbl
        self.files_table.installEventFilter(self)
        self._reposition_logo_overlay()

    def _reposition_logo_overlay(self):
        """Move the logo to the bottom-left corner of the files table (10 px margin)."""
        if self._logo_overlay is None:
            return
        tbl = self.files_table
        lbl = self._logo_overlay
        y = tbl.height() - lbl.height() - 10
        lbl.move(10, max(0, y))

    def _on_compound_hover(self, item: QTreeWidgetItem, column: int) -> None:
        """Show the structure popup for the compound row the mouse entered."""
        if not self.isActiveWindow():
            # Only show hover popup when the main window has focus
            return
        compound_data = item.data(0, Qt.ItemDataRole.UserRole)
        if not compound_data:
            # Group header rows carry no compound data
            self.structure_popup.hide()
            return
        self.structure_popup.show_for_compound(compound_data, QCursor.pos())

    def update_compounds_table(self):
        """Update the compounds table with loaded data"""
        self.compounds_table.clear()
        self.compounds_table.setHeaderLabels(["Name", "Retention Time", "Type", "Quantification", "Formula", "Mass", "Common Adducts"])
        compounds_data = self.compound_manager.get_compounds_data()

        if hasattr(self, "_compounds_placeholder"):
            self._compounds_placeholder.setVisible(compounds_data.empty)

        if compounds_data.empty:
            return

        # Build group structure
        group_dict = {}  # {group_name: [compound_indices]}
        ungrouped_compounds = []  # Compounds without a group

        for idx, (_, compound) in enumerate(compounds_data.iterrows()):
            compound_name = compound["Name"]
            groups_str = compound.get("Group", "")

            # Parse groups (semicolon-separated)
            # Check if group is not None/NaN and has actual content
            if pd.notna(groups_str) and groups_str is not None and str(groups_str).strip() and str(groups_str).strip().lower() != "none":
                groups = [g.strip() for g in str(groups_str).split(";") if g.strip()]
                if groups:  # Only process if we have actual groups after splitting
                    for group in groups:
                        if group not in group_dict:
                            group_dict[group] = []
                        group_dict[group].append(idx)
                else:
                    # No valid groups after parsing - add to ungrouped
                    ungrouped_compounds.append(idx)
            else:
                # No group specified - add to ungrouped
                ungrouped_compounds.append(idx)

        # Add grouped compounds (sorted by group name)
        for group_name in sorted(group_dict.keys()):
            compound_indices = group_dict[group_name]

            # Add group header as a top-level item
            group_item = QTreeWidgetItem(self.compounds_table)
            group_item.setText(0, group_name)
            group_item.setFont(0, QFont("", -1, QFont.Weight.Bold))
            group_item.setBackground(0, QColor(230, 230, 230))
            group_item.setBackground(1, QColor(230, 230, 230))
            group_item.setBackground(2, QColor(230, 230, 230))
            group_item.setBackground(3, QColor(230, 230, 230))
            group_item.setBackground(4, QColor(230, 230, 230))
            group_item.setBackground(5, QColor(230, 230, 230))
            group_item.setBackground(6, QColor(230, 230, 230))
            group_item.setExpanded(True)  # Expand by default

            # Add compounds under this group
            for idx in compound_indices:
                compound = compounds_data.iloc[idx]
                self._add_compound_row_tree(group_item, compound)

        # Add ungrouped compounds in their own group
        if ungrouped_compounds:
            # Add empty group header for ungrouped compounds
            group_item = QTreeWidgetItem(self.compounds_table)
            group_item.setText(0, " - no group")
            group_item.setFont(0, QFont("", -1, QFont.Weight.Bold))
            group_item.setBackground(0, QColor(245, 245, 245))
            group_item.setBackground(1, QColor(245, 245, 245))
            group_item.setBackground(2, QColor(245, 245, 245))
            group_item.setBackground(3, QColor(245, 245, 245))
            group_item.setBackground(4, QColor(245, 245, 245))
            group_item.setBackground(5, QColor(245, 245, 245))
            group_item.setBackground(6, QColor(245, 245, 245))
            group_item.setExpanded(True)  # Expand by default

            # Add ungrouped compounds
            for idx in ungrouped_compounds:
                compound = compounds_data.iloc[idx]
                self._add_compound_row_tree(group_item, compound)

    def _add_compound_row(self, row_idx, compound, indent=False):
        """Helper method to add a compound row to the table (legacy compatibility)"""
        compound_name = compound["Name"]

        # Compound name (with indent if under a group)
        display_name = f"  {compound_name}" if indent else compound_name
        name_item = QTableWidgetItem(display_name)
        name_item.setData(Qt.ItemDataRole.UserRole, compound.to_dict())
        self.compounds_table.setItem(row_idx, 0, name_item)

        # Retention time info
        rt_text = ""
        if "RT_minutes" in compound and pd.notna(compound["RT_minutes"]):
            avg_rt = compound["RT_minutes"]
            rt_text = f"{avg_rt:.1f} min"
        elif compound.get("RT_start_min") and compound.get("RT_end_min"):
            rt_start = compound["RT_start_min"]
            rt_end = compound["RT_end_min"]
            # Check if using default range (0-100 min)
            if rt_start == 0.0 and rt_end == 100.0:
                pass
            else:
                rt_text = f"{rt_start:.1f}-{rt_end:.1f} min"

        rt_item = QTableWidgetItem(rt_text)
        self.compounds_table.setItem(row_idx, 1, rt_item)

        # Compound type
        compound_type = compound.get("compound_type", "formula")
        type_display = {
            "formula": "Formula",
            "mass": "Mass",
            "mz_only": "Adduct",
        }.get(compound_type, compound_type)

        type_item = QTableWidgetItem(type_display)
        self.compounds_table.setItem(row_idx, 2, type_item)

    def _add_compound_row_tree(self, parent_item, compound):
        """Helper method to add a compound row to the tree widget"""
        compound_name = compound["Name"]

        # Create tree item as child of parent
        item = QTreeWidgetItem(parent_item)
        item.setText(0, compound_name)
        item.setData(0, Qt.ItemDataRole.UserRole, compound.to_dict())

        # Retention time info
        rt_text = ""
        if "RT_minutes" in compound and pd.notna(compound["RT_minutes"]):
            avg_rt = compound["RT_minutes"]
            rt_text = f"{avg_rt:.1f} min"
        elif compound.get("RT_start_min") and compound.get("RT_end_min"):
            rt_start = compound["RT_start_min"]
            rt_end = compound["RT_end_min"]
            # Check if using default range (0-100 min)
            if rt_start == 0.0 and rt_end == 100.0:
                pass
            else:
                rt_text = f"{rt_start:.1f}-{rt_end:.1f} min"

        item.setText(1, rt_text)

        # Compound type
        compound_type = compound.get("compound_type", "formula")
        type_display = {
            "formula": "Formula",
            "mass": "Mass",
            "mz_only": "Adduct",
        }.get(compound_type, compound_type)

        item.setText(2, type_display)

        # Quantification range
        quant_text = self._get_quant_range_for_compound(compound["Name"])
        item.setText(3, quant_text)

        # Chemical formula
        formula = compound.get("ChemicalFormula", "")
        if formula is None or (isinstance(formula, float) and pd.isna(formula)):
            formula = ""
        item.setText(4, str(formula).strip())

        # Mass
        mass = compound.get("Mass", "")
        if mass is None or (isinstance(mass, float) and pd.isna(mass)):
            mass_text = ""
        else:
            try:
                mass_text = f"{float(mass):.4f}"
            except (TypeError, ValueError):
                mass_text = str(mass).strip()
        item.setText(5, mass_text)

        # Common adducts
        adducts = compound.get("Common_adducts", "")
        if adducts is None or (isinstance(adducts, float) and pd.isna(adducts)):
            adducts_text = ""
        elif isinstance(adducts, list):
            adducts_text = ", ".join(str(a) for a in adducts)
        else:
            adducts_text = str(adducts).strip()
        item.setText(6, adducts_text)

    def show_files_context_menu(self, position):
        """Show context menu for file operations"""
        item = self.files_table.itemAt(position)
        if item is None:
            return

        # Get the selected row
        row = item.row()

        # Create context menu
        menu = QMenu(self)

        # Get file information for display
        file_name = "Unknown"
        filepath = None
        if self.files_table.columnCount() > 0:
            name_item = self.files_table.item(row, 0)
            if name_item:
                file_name = name_item.text()
            # Find the Filepath column
            for col in range(self.files_table.columnCount()):
                header = self.files_table.horizontalHeaderItem(col)
                if header and header.text() == "Filepath":
                    fp_item = self.files_table.item(row, col)
                    if fp_item:
                        filepath = fp_item.text()
                    break

        # Add explore file action
        if filepath and os.path.exists(filepath):
            explore_action = QAction(f"Explore File", self)
            explore_action.triggered.connect(lambda checked, fp=filepath: self._open_file_explorer(fp))
            menu.addAction(explore_action)
            menu.addSeparator()

        # Add remove file action
        remove_action = QAction(f"Remove File", self)
        remove_action.triggered.connect(lambda checked, r=row: self.remove_file_at_row(r))
        menu.addAction(remove_action)

        # Show menu at cursor position
        menu.exec(self.files_table.mapToGlobal(position))

    def _open_file_explorer(self, filepath: str):
        """Open the mzML file explorer window for the given file."""
        win = MzMLFileExplorerWindow(filepath, self.file_manager, defaults=self.eic_defaults, parent=None)
        win.show()
        if not hasattr(self, "_file_explorer_windows"):
            self._file_explorer_windows = []
        self._file_explorer_windows.append(win)
        win.destroyed.connect(lambda _, w=win: self._file_explorer_windows.remove(w) if hasattr(self, "_file_explorer_windows") and w in self._file_explorer_windows else None)

    def remove_file_at_row(self, row):
        """Remove a file at the specified row"""
        if row < 0 or row >= self.files_table.rowCount():
            return

        # Get file information for confirmation
        file_name = "Unknown"
        if self.files_table.columnCount() > 0:
            name_item = self.files_table.item(row, 0)
            if name_item:
                file_name = name_item.text()

        # Confirm removal
        reply = QMessageBox.question(
            self,
            "Remove File",
            f"Are you sure you want to remove the file '{file_name}' from the list?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No,
        )

        if reply == QMessageBox.StandardButton.Yes:
            try:
                # Get the current files data
                files_data = self.file_manager.get_files_data()

                if not files_data.empty and row < len(files_data):
                    # Remove the file from the file manager
                    files_data_copy = files_data.copy()
                    files_data_copy = files_data_copy.drop(files_data_copy.index[row]).reset_index(drop=True)

                    # Update the file manager with the new data
                    self.file_manager.files_data = files_data_copy

                    # Update the UI
                    self.update_files_table()

                    # Update status
                    remaining_files = len(self.file_manager.get_files_data())
                    self.statusBar().showMessage(f"File removed. Remaining: {remaining_files} files")

            except Exception as e:
                QMessageBox.critical(self, "Error", f"Failed to remove file: {str(e)}")

    def show_compound_context_menu(self, position):
        """Show context menu for compound adducts"""
        item = self.compounds_table.itemAt(position)
        if item is None:
            return

        # Get the compound data from the item
        compound_data = item.data(0, Qt.ItemDataRole.UserRole)

        # Skip if this is a group header (no UserRole data)
        if not compound_data:
            return

        compound_name = compound_data["Name"]

        # Create context menu
        menu = QMenu(self)
        menu.setTitle(f"Adducts for {compound_name}")

        # --- non-clickable info header (formula / mass) ---
        formula = compound_data.get("ChemicalFormula")
        mass = compound_data.get("Mass")
        # Treat NaN / None / empty string as absent
        formula_ok = formula is not None and not (isinstance(formula, float) and pd.isna(formula)) and str(formula).strip()
        mass_ok = mass is not None and not (isinstance(mass, float) and pd.isna(mass)) and str(mass).strip()
        info_lines = [f"<b>{compound_name}</b>"]
        if formula_ok:
            info_lines.append(f"Formula: {formula}")
            if not mass_ok:
                # calculate mass from formula for display
                try:
                    from .utils import calculate_molecular_mass

                    calc_mass = calculate_molecular_mass(str(formula).strip())
                    info_lines.append(f"Mass: {calc_mass:.4f} Da")
                except Exception:
                    pass
        if mass_ok:
            info_lines.append(f"Mass: {mass} Da")
        if len(info_lines) > 1:  # only show when there is something beyond the name
            info_label = QLabel("<br>".join(info_lines))
            info_label.setStyleSheet("padding: 4px 8px; color: #333; background: transparent;")
            info_label.setTextFormat(Qt.TextFormat.RichText)
            info_widget_action = QWidgetAction(self)
            info_widget_action.setDefaultWidget(info_label)
            info_widget_action.setEnabled(False)
            menu.addAction(info_widget_action)
            menu.addSeparator()

        # Get categorized adducts
        adducts_info = self.compound_manager.get_compound_adducts_categorized(compound_name)
        can_calculate = self.compound_manager.can_calculate_adducts_from_formula(compound_name)

        # Add specified adducts
        specified_adducts = adducts_info["specified"]
        if specified_adducts:
            if can_calculate:
                # Show specified adducts at top for compounds with formula/mass
                for adduct in specified_adducts:
                    self._add_adduct_action(menu, compound_data, adduct, specified=True)

                # Add separator
                if adducts_info["remaining"]:
                    menu.addSeparator()

                # Add remaining adducts
                for adduct in adducts_info["remaining"]:
                    self._add_adduct_action(menu, compound_data, adduct, specified=False)
            else:
                # For m/z only compounds, show only the specified adducts
                for adduct in specified_adducts:
                    self._add_adduct_action(menu, compound_data, adduct, specified=True)
        elif can_calculate:
            # No adducts specified but can calculate from formula - show all possible
            all_adducts = self.compound_manager.get_all_available_adducts()
            for adduct in all_adducts:
                self._add_adduct_action(menu, compound_data, adduct, specified=False)
        else:
            # No adducts and can't calculate - show message
            no_adducts_action = QAction("No adducts available", self)
            no_adducts_action.setEnabled(False)
            menu.addAction(no_adducts_action)

        # Add isotopolog submenu (requires at least one adduct option)
        separator_action = None
        if menu.actions():
            separator_action = menu.addSeparator()

        isotopolog_added = self._add_isotopolog_menu(menu, compound_data, adducts_info, can_calculate)

        if not isotopolog_added and separator_action is not None:
            menu.removeAction(separator_action)

        # Add separator before multi-adduct options
        if menu.actions():
            menu.addSeparator()

        # Add multi-adduct options
        self._add_multi_adduct_actions(menu, compound_data, adducts_info, can_calculate)

        # Show menu at cursor position
        if menu.actions():
            menu.exec(self.compounds_table.mapToGlobal(position))

    def _add_adduct_action(self, menu, compound_data, adduct, specified=True):
        """Add an adduct action to the context menu"""
        # Get pre-calculated data for display
        compound_name = compound_data["Name"]
        precalc_data = self.compound_manager.get_precalculated_data(compound_name, adduct)

        if precalc_data:
            display_name = precalc_data["display_name"]
            mz_value = precalc_data["mz"]
            polarity = precalc_data["polarity"]

            if mz_value is not None:
                action_text = f"{display_name} (m/z: {mz_value:.4f})"
            else:
                action_text = f"{display_name} (m/z: calculation failed)"
        else:
            # Try to calculate on the fly
            display_name = self.compound_manager.get_adduct_display_name(compound_name, adduct)
            mz_value = self.compound_manager.calculate_compound_mz(compound_name, adduct)
            polarity = self.compound_manager._determine_polarity(adduct)

            if mz_value is not None:
                action_text = f"{display_name} (m/z: {mz_value:.4f})"
            else:
                action_text = f"{display_name} (m/z: not calculated)"

        if specified:
            # Use QWidgetAction with an HTML-bold label so bold text is guaranteed
            # to render on all platforms (QAction.setFont is ignored by the native
            # Windows menu renderer).
            label = QLabel(f"<b>{action_text}</b>")
            label.setTextFormat(Qt.TextFormat.RichText)
            bold_font = label.font()
            bold_font.setBold(True)
            label.setFont(bold_font)
            label.setContentsMargins(20, 3, 8, 3)
            label.setAutoFillBackground(False)
            # Pass mouse events through to the menu so that clicking the label
            # correctly triggers the QWidgetAction (labels swallow mouse events
            # by default, which prevents the triggered signal from firing).
            label.setAttribute(Qt.WidgetAttribute.WA_TransparentForMouseEvents, True)
            waction = QWidgetAction(self)
            waction.setDefaultWidget(label)
            waction.triggered.connect(lambda checked, c=compound_data, a=adduct, m=mz_value, p=polarity: self.show_eic_window(c, a, m, p))
            menu.addAction(waction)
        else:
            action = QAction(action_text, self)
            action.triggered.connect(lambda checked, c=compound_data, a=adduct, m=mz_value, p=polarity: self.show_eic_window(c, a, m, p))
            menu.addAction(action)

    def _add_isotopolog_menu(self, menu, compound_data, adducts_info, can_calculate) -> bool:
        """Add an isotopolog submenu allowing selection of labeled variants."""
        compound_name = compound_data["Name"]

        # Gather available adducts (specified first, then remaining if calculable)
        available_adducts = list(adducts_info["specified"])
        if can_calculate:
            available_adducts.extend(adducts_info["remaining"])

        # Remove duplicates while preserving order
        seen_adducts = set()
        ordered_adducts = []
        for adduct in available_adducts:
            if adduct not in seen_adducts:
                ordered_adducts.append(adduct)
                seen_adducts.add(adduct)

        if not ordered_adducts:
            return False

        element_map = self.compound_manager.get_isotopolog_elements(compound_name)
        if not element_map:
            return False

        isotopolog_menu = None
        any_added = False

        custom_definitions = self._parse_custom_isotopolog_definitions(compound_data)
        if custom_definitions:
            if isotopolog_menu is None:
                isotopolog_menu = menu.addMenu("Isotopologs")

            custom_menu = isotopolog_menu.addMenu("Custom")
            for custom_name, isotopolog_formula in custom_definitions.items():
                custom_entry_menu = custom_menu.addMenu(custom_name)
                custom_has_action = False

                for adduct in ordered_adducts:
                    custom_mz = self.compound_manager.calculate_custom_isotopolog_mz(compound_name, adduct, isotopolog_formula)
                    adduct_display = self.compound_manager.get_adduct_display_name(compound_name, adduct)

                    if custom_mz is None:
                        action = QAction(f"{adduct_display} (m/z unavailable)", self)
                        action.setEnabled(False)
                        custom_entry_menu.addAction(action)
                        continue

                    action = QAction(f"{adduct_display} (m/z: {custom_mz:.4f})", self)

                    precalc = self.compound_manager.get_precalculated_data(compound_name, adduct)
                    polarity = precalc.get("polarity") if precalc else None
                    if polarity is None:
                        polarity = self.compound_manager._determine_polarity(adduct)

                    adduct_label = f"{adduct_display} {custom_name}"
                    action.triggered.connect(lambda checked, c=compound_data, label=adduct_label, mz=custom_mz, pol=polarity: self.show_eic_window(c, label, mz, pol))
                    custom_entry_menu.addAction(action)
                    custom_has_action = True

                if not custom_has_action:
                    custom_entry_menu.menuAction().setEnabled(False)
                else:
                    any_added = True

        for element, total_count in element_map.items():
            isotope_label = self.compound_manager.get_isotope_label(element)
            if not isotope_label:
                continue

            isotopolog_variants = self._build_isotopolog_variants(element, isotope_label, total_count)
            if not isotopolog_variants:
                continue

            element_menu = None

            for display_label, count in isotopolog_variants:
                count_actions = []

                for adduct in ordered_adducts:
                    isotopolog_mz = self.compound_manager.calculate_isotopolog_mz(compound_name, adduct, element, count)

                    adduct_display = self.compound_manager.get_adduct_display_name(compound_name, adduct)

                    if isotopolog_mz is None:
                        action_text = f"{adduct_display} (m/z unavailable)"
                        action = QAction(action_text, self)
                        action.setEnabled(False)
                    else:
                        action_text = f"{adduct_display} (m/z: {isotopolog_mz:.4f})"
                        action = QAction(action_text, self)

                        precalc = self.compound_manager.get_precalculated_data(compound_name, adduct)
                        polarity = None
                        if precalc:
                            polarity = precalc.get("polarity")
                        if polarity is None:
                            polarity = self.compound_manager._determine_polarity(adduct)

                        adduct_label = f"{adduct_display} {display_label}"

                        action.triggered.connect(lambda checked, c=compound_data, label=adduct_label, mz=isotopolog_mz, pol=polarity: self.show_eic_window(c, label, mz, pol))

                    count_actions.append(action)

                if not count_actions:
                    continue

                if isotopolog_menu is None:
                    isotopolog_menu = menu.addMenu("Isotopologs")

                if element_menu is None:
                    element_menu = isotopolog_menu.addMenu(element)

                count_menu = element_menu.addMenu(display_label)
                for action in count_actions:
                    count_menu.addAction(action)

                any_added = True

        if not any_added and isotopolog_menu is not None:
            isotopolog_menu.menuAction().setEnabled(False)

        return any_added

    def _build_isotopolog_variants(self, element, isotope_label, total_count):
        """Build display labels and counts for isotopolog menu entries."""
        variants = []

        counts = self.compound_manager.get_isotopolog_counts(element, total_count)
        for count in counts:
            variants.append((f"[{isotope_label}]{count}", count))

        # Additional C-specific shortcuts requested by users.
        if element == "C" and total_count is not None:
            try:
                total_c = int(total_count)
            except (TypeError, ValueError):
                total_c = None

            if total_c and total_c > 0:
                if total_c - 1 > 0:
                    variants.append((f"-1 C ([{isotope_label}]{total_c - 1})", total_c - 1))
                if total_c - 2 > 0:
                    variants.append((f"-2 C ([{isotope_label}]{total_c - 2})", total_c - 2))

                variants.append((f"total C+1 ([{isotope_label}]{total_c + 1})", total_c + 1))
                variants.append((f"total C+2 ([{isotope_label}]{total_c + 2})", total_c + 2))

        # Deduplicate by (label, count) while preserving order.
        seen = set()
        deduped = []
        for label, count in variants:
            key = (label, count)
            if key in seen:
                continue
            seen.add(key)
            deduped.append((label, count))

        return deduped

    def _parse_custom_isotopolog_definitions(self, compound_data):
        """Parse optional Isotopologs JSON string from a compound row."""
        raw_value = compound_data.get("Isotopologs", "")
        if raw_value is None or (isinstance(raw_value, float) and pd.isna(raw_value)):
            return {}

        if not isinstance(raw_value, str):
            return {}

        raw_value = raw_value.strip()
        if not raw_value:
            return {}

        try:
            parsed = json.loads(raw_value)
        except Exception:
            return {}

        if not isinstance(parsed, dict):
            return {}

        result = {}
        for key, value in parsed.items():
            if not isinstance(key, str) or not isinstance(value, str):
                continue
            name = key.strip()
            formula = value.strip()
            if not name or not formula:
                continue
            result[name] = formula

        return result

    def _add_multi_adduct_actions(self, menu, compound_data, adducts_info, can_calculate):
        """Add multi-adduct window actions to the context menu"""
        compound_name = compound_data["Name"]
        compound_type = compound_data.get("compound_type", "formula")

        # Compounds with formula or mass can calculate the full adduct list
        can_show_all_adducts = compound_type in ["formula", "mass"]

        # Option 1: Show predefined adducts in multi-EIC window
        specified_adducts = adducts_info["specified"]

        valid_predefined_adducts = []
        for adduct in specified_adducts:
            precalc = self.compound_manager.get_precalculated_data(compound_name, adduct)
            mz_value = None
            if precalc:
                mz_value = precalc.get("mz")
            if mz_value is None:
                mz_value = self.compound_manager.calculate_compound_mz(compound_name, adduct)
            if mz_value is not None:
                valid_predefined_adducts.append(adduct)

        if valid_predefined_adducts:
            predefined_action = QAction("📊 Show Predefined Adducts (Multi-EIC)", self)
            predefined_action.setEnabled(True)
            predefined_action.triggered.connect(lambda checked, c=compound_data: self.show_multi_adduct_window(c, show_predefined_only=True))
            menu.addAction(predefined_action)
        elif specified_adducts:
            # Provide contextual feedback when adducts exist but no m/z can be derived
            predefined_action = QAction("📊 Show Predefined Adducts (Multi-EIC) - no valid m/z", self)
            predefined_action.setEnabled(False)
            predefined_action.setToolTip("All predefined adducts for this compound are missing calculable m/z values.")
            menu.addAction(predefined_action)

        # Option 2: Show all possible adducts in multi-EIC window
        # This option is only shown if we can calculate adducts (formula or mass available)
        all_adducts_action = None
        if can_calculate and can_show_all_adducts:
            all_adducts_action = QAction("📊 Show All Adducts (Multi-EIC)", self)
            all_adducts_action.triggered.connect(lambda checked, c=compound_data: self.show_multi_adduct_window(c, show_predefined_only=False))
            menu.addAction(all_adducts_action)

        # If no multi-adduct options are available for m/z only compounds, show informational message
        if not can_show_all_adducts and not valid_predefined_adducts:
            info_action = QAction(
                "ℹ️ Multi-adduct view requires formula, mass, or valid m/z adducts",
                self,
            )
            info_action.setEnabled(False)
            menu.addAction(info_action)

    def show_multi_adduct_window(self, compound, show_predefined_only=True):
        """Show multi-adduct EIC window"""
        if self.file_manager.get_files_data().empty:
            QMessageBox.warning(self, "Warning", "No files loaded!")
            return

        try:
            compound_name = compound["Name"]

            # Prepare adducts data
            adducts_data = []

            if show_predefined_only:
                # Get only predefined adducts
                specified_adducts = self.compound_manager.get_compound_adducts(compound_name)
                for adduct in specified_adducts:
                    precalc = self.compound_manager.get_precalculated_data(compound_name, adduct)
                    mz_value = None
                    polarity = None

                    if precalc:
                        mz_value = precalc.get("mz")
                        polarity = precalc.get("polarity")

                    if mz_value is None:
                        mz_value = self.compound_manager.calculate_compound_mz(compound_name, adduct)

                    if polarity is None:
                        polarity = self.compound_manager._determine_polarity(adduct)

                    if mz_value is not None:
                        adducts_data.append((adduct, mz_value, polarity))
            else:
                # Get all possible adducts
                if self.compound_manager.can_calculate_adducts_from_formula(compound_name):
                    all_adducts = self.compound_manager.get_all_available_adducts()
                    for adduct in all_adducts:
                        mz_value = self.compound_manager.calculate_compound_mz(compound_name, adduct)
                        polarity = self.compound_manager._determine_polarity(adduct)
                        if mz_value is not None:  # Only add adducts with valid m/z
                            adducts_data.append((adduct, mz_value, polarity))
                else:
                    # For m/z only compounds, fall back to predefined
                    specified_adducts = self.compound_manager.get_compound_adducts(compound_name)
                    for adduct in specified_adducts:
                        mz_value = self.compound_manager.calculate_compound_mz(compound_name, adduct)
                        polarity = self.compound_manager._determine_polarity(adduct)
                        if mz_value is not None:
                            adducts_data.append((adduct, mz_value, polarity))

            if not adducts_data:
                QMessageBox.information(self, "Information", "No adducts available for this compound.")
                return

            # Create multi-adduct window
            multi_window = MultiAdductWindow(
                compound,
                adducts_data,
                self.file_manager,
                defaults=self.eic_defaults,
                show_predefined_only=show_predefined_only,
                parent=None,
            )

            # Show the window
            multi_window.show()
            multi_window.raise_()
            multi_window.activateWindow()

            # Keep reference to prevent garbage collection
            self.eic_windows.append(multi_window)

            # Remove from list when window is closed
            def on_window_closed():
                if multi_window in self.eic_windows:
                    self.eic_windows.remove(multi_window)

            multi_window.destroyed.connect(on_window_closed)

        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to create multi-adduct window: {str(e)}")
            import traceback

            traceback.print_exc()

    def filter_compounds(self):
        """Filter compounds based on the filter text"""
        filter_text = self.compound_filter.text().strip()

        if not filter_text:
            # Show all compounds if filter is empty
            iterator = QTreeWidgetItemIterator(self.compounds_table)
            while iterator.value():
                item = iterator.value()
                item.setHidden(False)
                iterator += 1
            return

        # Parse filter text
        filter_type, filter_params = self._parse_filter_text(filter_text)

        # Iterate through all items in the tree
        iterator = QTreeWidgetItemIterator(self.compounds_table)
        while iterator.value():
            item = iterator.value()
            compound_data = item.data(0, Qt.ItemDataRole.UserRole)

            # This is a group header if there's no compound data
            if not compound_data:
                # Keep group headers visible, but they'll be hidden if all children are hidden
                iterator += 1
                continue

            compound_name = compound_data["Name"]
            show_compound = False

            if filter_type == "mz":
                # Check if any adducts have m/z in range
                min_mz, max_mz = filter_params
                adducts = compound_data.get("Common_adducts", "")

                # Build explicit adduct list from Common_adducts
                if isinstance(adducts, list):
                    adduct_list = [a.strip() for a in adducts if a and str(a).strip()]
                elif isinstance(adducts, str) and adducts.strip():
                    adduct_list = [a.strip() for a in adducts.split(",") if a.strip()]
                else:
                    adduct_list = []

                # For formula/mass compounds also consider all adducts from the adducts table
                compound_type = compound_data.get("compound_type", "formula")
                if compound_type in ("formula", "mass"):
                    adducts_df = self.compound_manager.get_adducts_data()
                    if not adducts_df.empty and "Adduct" in adducts_df.columns:
                        all_table_adducts = adducts_df["Adduct"].dropna().tolist()
                        # Append adducts not already in the explicit list
                        explicit_set = set(adduct_list)
                        adduct_list = adduct_list + [a for a in all_table_adducts if a not in explicit_set]

                for adduct in adduct_list:
                    try:
                        # Use pre-calculated m/z value if available
                        precalc_data = self.compound_manager.get_precalculated_data(compound_name, adduct)

                        if precalc_data and precalc_data["mz"] is not None:
                            mz_value = precalc_data["mz"]
                        else:
                            # Fallback to calculation if pre-calculated data not available
                            mz_value = self.compound_manager.calculate_compound_mz(compound_name, adduct)

                        if mz_value is not None and min_mz <= mz_value <= max_mz:
                            show_compound = True
                            break
                    except:
                        continue

            elif filter_type == "rt":
                # Check if RT is in range
                min_rt, max_rt = filter_params
                if "RT_minutes" in compound_data and pd.notna(compound_data["RT_minutes"]):
                    rt_value = float(compound_data["RT_minutes"])
                    show_compound = min_rt <= rt_value <= max_rt
                elif compound_data.get("RT_start_min") and compound_data.get("RT_end_min"):
                    # Use average of RT window if RT_minutes not available
                    avg_rt = (float(compound_data["RT_start_min"]) + float(compound_data["RT_end_min"])) / 2
                    show_compound = min_rt <= avg_rt <= max_rt

            elif filter_type == "smarts":
                # SMARTS substructure match via RDKit
                smarts_mol, smarts_err = filter_params
                if smarts_mol is None:
                    # Invalid SMARTS — hide everything so the user notices
                    show_compound = False
                else:
                    smiles = compound_data.get("SMILES") or compound_data.get("smiles")
                    if smiles and str(smiles).strip() and str(smiles).strip().lower() not in ("nan", "none"):
                        try:
                            from rdkit import Chem

                            mol = Chem.MolFromSmiles(str(smiles).strip())
                            show_compound = mol is not None and mol.HasSubstructMatch(smarts_mol)
                        except Exception:
                            show_compound = False
                    else:
                        show_compound = False

            elif filter_type == "name":
                # Regex search on compound name
                import re

                try:
                    pattern = re.compile(filter_params, re.IGNORECASE)
                    show_compound = bool(pattern.search(compound_name))
                except re.error:
                    # If regex is invalid, fall back to simple string search
                    show_compound = filter_params.lower() in compound_name.lower()

            # Show/hide compound based on filter result
            item.setHidden(not show_compound)
            iterator += 1

        # Hide group headers if all their children are hidden
        root = self.compounds_table.invisibleRootItem()
        for i in range(root.childCount()):
            group_item = root.child(i)
            all_children_hidden = True
            for j in range(group_item.childCount()):
                if not group_item.child(j).isHidden():
                    all_children_hidden = False
                    break
            group_item.setHidden(all_children_hidden)

    def _parse_filter_text(self, filter_text):
        """
        Parse filter text to determine filter type and parameters.

        Returns:
            tuple: (filter_type, parameters)
                - 'mz': parameters = (min_mz, max_mz)
                - 'rt': parameters = (min_rt, max_rt)
                - 'name': parameters = regex_pattern
        """
        import re

        # Check for mz filter: "mz 100-200" or "mz 100 - 200"
        mz_match = re.match(r"mz\s+(\d+(?:\.\d+)?)\s*-\s*(\d+(?:\.\d+)?)", filter_text, re.IGNORECASE)
        if mz_match:
            min_mz = float(mz_match.group(1))
            max_mz = float(mz_match.group(2))
            return "mz", (min_mz, max_mz)

        # Check for rt filter: "rt 5-10" or "rt 5 - 10"
        rt_match = re.match(r"rt\s+(\d+(?:\.\d+)?)\s*-\s*(\d+(?:\.\d+)?)", filter_text, re.IGNORECASE)
        if rt_match:
            min_rt = float(rt_match.group(1))
            max_rt = float(rt_match.group(2))
            return "rt", (min_rt, max_rt)

        # Check for SMARTS filter: "SMARTS:<pattern>"
        smarts_match = re.match(r"smarts?:(.+)", filter_text, re.IGNORECASE)
        if smarts_match:
            smarts_str = smarts_match.group(1).strip()
            try:
                from rdkit import Chem

                smarts_mol = Chem.MolFromSmarts(smarts_str)
            except Exception:
                smarts_mol = None
            return "smarts", (smarts_mol, smarts_str)

        # Otherwise, treat as name regex pattern
        return "name", filter_text

    def show_custom_eic_dialog(self):
        """Open the 'Show Custom EIC Trace' dialog."""
        # Always start from the full template list so the combo is never empty.
        # Append any extra adducts from a loaded compounds file that are not
        # already covered by the template.
        base_df = pd.DataFrame(ADDUCTS_TEMPLATE_DATA)
        loaded = self.compound_manager.adducts_data
        if not loaded.empty:
            template_names = set(base_df["Adduct"].tolist())
            extra = loaded[~loaded["Adduct"].isin(template_names)]
            if not extra.empty:
                base_df = pd.concat([base_df, extra], ignore_index=True)
        adducts_data = base_df

        dialog = CustomEICDialog(adducts_data, ppm_default=self.eic_defaults["mz_tolerance_ppm"], parent=self)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            compound_data, adduct, mz_value, polarity, ppm = dialog.get_result()
            self.show_eic_window(compound_data, adduct, mz_value=mz_value, polarity=polarity, ppm_override=ppm)

    def show_eic_window(self, compound, adduct, mz_value=None, polarity=None, ppm_override=None):
        """Show EIC window for the selected compound and adduct"""
        if self.file_manager.get_files_data().empty:
            QMessageBox.warning(self, "Warning", "No files loaded!")
            return

        try:
            # Create EIC window as independent window (no parent)
            effective_defaults = self.eic_defaults
            if ppm_override is not None:
                effective_defaults = dict(self.eic_defaults)
                effective_defaults["mz_tolerance_ppm"] = ppm_override
            eic_window = EICWindow(
                compound,
                adduct,
                self.file_manager,
                mz_value=mz_value,
                polarity=polarity,
                defaults=effective_defaults,  # Pass the defaults
                parent=None,  # Make it independent
                integration_callback=self.record_peak_integration,  # Pass integration callback
                settings_callback=self._on_eic_settings_changed,  # Persist control changes
                adducts_data=self.compound_manager.adducts_data,  # For fragment annotation
            )

            # Show the window
            eic_window.show()
            eic_window.raise_()  # Bring to front
            eic_window.activateWindow()  # Make it the active window

            # Keep reference to prevent garbage collection
            self.eic_windows.append(eic_window)

            # Remove from list when window is closed
            def on_window_closed():
                if eic_window in self.eic_windows:
                    self.eic_windows.remove(eic_window)

            eic_window.destroyed.connect(on_window_closed)

        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to create EIC window: {str(e)}")
            import traceback

            traceback.print_exc()

    def _open_spectrum_comparator(self):
        """Open an empty Spectrum Comparator window."""
        if not hasattr(self, "_comparator_windows"):
            self._comparator_windows = []
        win = USISpectrumComparisonWindow(
            file_manager=self.file_manager,
        )
        self._comparator_windows.append(win)
        win.destroyed.connect(lambda _, w=win: self._comparator_windows.remove(w) if hasattr(self, "_comparator_windows") and w in self._comparator_windows else None)
        win.show()
        win.raise_()

    def create_menu_bar(self):
        """Create the menu bar"""
        menubar = self.menuBar()

        # File menu
        file_menu = menubar.addMenu("File")

        # Load Files action
        load_files_action = QAction("Load mzML File Table", self)
        load_files_action.triggered.connect(self.load_files)
        file_menu.addAction(load_files_action)

        # Load Compounds action
        load_compounds_action = QAction("Load Compounds Table", self)
        load_compounds_action.triggered.connect(self.load_compounds)
        file_menu.addAction(load_compounds_action)

        file_menu.addSeparator()

        # Clear Files action
        clear_files_action = QAction("Clear mzML Files", self)
        clear_files_action.triggered.connect(self.clear_files)
        file_menu.addAction(clear_files_action)

        # Clear Compounds action
        clear_compounds_action = QAction("Clear Compounds", self)
        clear_compounds_action.triggered.connect(self.clear_compounds)
        file_menu.addAction(clear_compounds_action)

        file_menu.addSeparator()

        # Generate Templates action
        generate_templates_action = QAction("Generate Excel Templates", self)
        generate_templates_action.triggered.connect(self.generate_templates)
        file_menu.addAction(generate_templates_action)

        file_menu.addSeparator()

        # Export Peak Integration Data action
        export_integration_action = QAction("Export Peak Integration Data", self)
        export_integration_action.triggered.connect(self.export_peak_integration_excel)
        file_menu.addAction(export_integration_action)

        # Generate R Code action
        generate_r_code_action = QAction("Generate R Code for Peak Data", self)
        generate_r_code_action.triggered.connect(self.generate_r_code)
        file_menu.addAction(generate_r_code_action)

        file_menu.addSeparator()

        # Exit action
        exit_action = QAction("Exit", self)
        exit_action.setShortcut("Ctrl+Q")
        exit_action.triggered.connect(self.close)
        file_menu.addAction(exit_action)

        # Options menu
        options_menu = menubar.addMenu("Options")

        # Unified Options action
        options_action = QAction("Options", self)
        options_action.setShortcut("Ctrl+P")
        options_action.triggered.connect(self.show_options_dialog)
        options_menu.addAction(options_action)

        # Tools menu
        tools_menu = menubar.addMenu("Tools")

        fetch_info_action = QAction("Fetch compound information", self)
        fetch_info_action.triggered.connect(self.fetch_compound_information)
        tools_menu.addAction(fetch_info_action)

        tools_menu.addSeparator()

        comparator_action = QAction("Spectrum Comparator", self)
        comparator_action.setToolTip("Open a window to compare two arbitrary MSMS spectra")
        comparator_action.triggered.connect(self._open_spectrum_comparator)
        tools_menu.addAction(comparator_action)

        # Help menu
        help_menu = menubar.addMenu("Help")

        # About action
        about_action = QAction("About", self)
        about_action.triggered.connect(self.show_about_dialog)
        help_menu.addAction(about_action)

    def show_about_dialog(self):
        """Show the about dialog"""
        from PyQt6.QtWidgets import QDialogButtonBox

        # Extract version from the project's pyproject.toml file
        try:
            toml_path = os.path.join(os.path.dirname(__file__), "..", "..", "pyproject.toml")
            with open(toml_path, "r") as f:
                project_data = toml.load(f)
                version = project_data.get("project", {}).get("version", "Unknown")
        except Exception:
            version = "Unknown"

        dlg = QDialog(self)
        dlg.setWindowTitle("About mzML Explorer")
        dlg.setMinimumWidth(400)
        dlg.setStyleSheet("background-color: white;")
        layout = QVBoxLayout(dlg)
        layout.setSpacing(12)
        layout.setContentsMargins(20, 20, 20, 20)

        # Logo
        if os.path.exists(_LOGO_PATH):
            from .window_game import SnakeWindow

            logo_lbl = QLabel()
            pix = QPixmap(_LOGO_PATH).scaledToWidth(260, Qt.TransformationMode.SmoothTransformation)
            logo_lbl.setPixmap(pix)
            logo_lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
            logo_lbl.setToolTip("Click to play Snake 🐍")

            def _open_snake(event, _dlg=dlg):
                _dlg.accept()
                snake_win = SnakeWindow(parent=None)
                snake_win.show()
                snake_win.raise_()
                # Keep reference on the main window so it isn't garbage-collected
                mw = self
                if not hasattr(mw, "_game_windows"):
                    mw._game_windows = []
                mw._game_windows.append(snake_win)
                snake_win.destroyed.connect(lambda: mw._game_windows.remove(snake_win) if snake_win in mw._game_windows else None)

            logo_lbl.mousePressEvent = _open_snake
            layout.addWidget(logo_lbl)

        # Info text
        text_lbl = QLabel(
            f"<b>mzML Explorer v{version}</b><br><br>"
            "A tool for visualizing LC-HRMS data from mzML files.<br><br>"
            "<b>Features:</b><br>"
            "\u2022 Load mzML files via Excel templates<br>"
            "\u2022 Extract ion chromatograms (EICs)<br>"
            "\u2022 Interactive plotting with zoom and pan<br>"
            "\u2022 Group-based color coding<br><br>"
            "Built with PyQt6 and pymzml.<br><br>"
            "\u00a9 2025 Plant-Microbe Metabolomics, BOKU University"
        )
        text_lbl.setWordWrap(True)
        text_lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(text_lbl)

        # OK button
        btn_box = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok)
        btn_box.accepted.connect(dlg.accept)
        layout.addWidget(btn_box)

        dlg.exec()

    def dragEnterEvent(self, event: QDragEnterEvent):
        """Handle drag enter events"""
        if event.mimeData().hasUrls():
            # Check if any of the files are Excel files
            for url in event.mimeData().urls():
                if url.isLocalFile():
                    file_path = url.toLocalFile()
                    if file_path.endswith((".xlsx", ".csv", ".tsv")):
                        event.acceptProposedAction()
                        return
        event.ignore()

    def dropEvent(self, event: QDropEvent):
        """Handle drop events"""
        files = []
        for url in event.mimeData().urls():
            if url.isLocalFile():
                file_path = url.toLocalFile()
                if file_path.endswith((".xlsx", ".csv", ".tsv")):
                    files.append(file_path)

        if files:
            # Determine drop location
            widget_under_mouse = self.childAt(event.position().toPoint())

            # Try to find which panel the drop occurred on
            files_panel_drop = False
            compounds_panel_drop = False

            # Walk up the widget hierarchy to find the target
            current_widget = widget_under_mouse
            while current_widget:
                if current_widget == self.files_table:
                    files_panel_drop = True
                    break
                elif current_widget == self.compounds_table:
                    compounds_panel_drop = True
                    break
                current_widget = current_widget.parent()

            # Process the first file
            file_path = files[0]

            try:
                if compounds_panel_drop:
                    # Load as compounds file
                    self.load_compounds_from_file(file_path)
                else:
                    # Default to loading as files list
                    self.load_files_from_file(file_path)

                event.acceptProposedAction()

            except Exception as e:
                QMessageBox.critical(self, "Error", f"Failed to load file: {str(e)}")

        event.ignore()

    def load_files_from_file(self, file_path):
        """Load files from a dropped file"""
        self._load_files_from_path(file_path, source="drag & drop")

    def load_compounds_from_file(self, file_path):
        """Load compounds from a dropped file"""
        try:
            file_ext = file_path.lower().split(".")[-1]

            if file_ext in ["csv", "tsv"]:
                # Use import dialog for CSV/TSV files
                self.load_compounds_from_csv_tsv(file_path)
            elif file_ext == "xlsx":
                # Load Excel file directly (existing functionality)
                self.load_compounds_from_excel(file_path)
            else:
                QMessageBox.warning(
                    self,
                    "Unsupported Format",
                    f"Unsupported file format: {file_ext}. Supported formats: xlsx, csv, tsv",
                )

        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to load compounds: {str(e)}")

    def load_compounds_from_csv_tsv(self, file_path):
        """Load compounds from CSV/TSV file using import dialog"""
        dialog = CompoundImportDialog(file_path, self)

        if dialog.exec() == QDialog.DialogCode.Accepted:
            import_data = dialog.get_import_data()
            if import_data is not None and not import_data.empty:
                # Load compounds
                self.compound_manager.load_compounds(import_data)
                self.update_compounds_table()

                compound_count = len(self.compound_manager.get_compounds_data())
                self.statusBar().showMessage(f"Compounds imported from {file_path}. Total: {compound_count} compounds")

                # Start monitoring the compound file for changes
                self.start_compound_file_monitoring(file_path)

    def load_compounds_from_excel(self, file_path):
        """Load compounds from Excel file (existing functionality)"""
        # Load all sheets from Excel file
        excel_data = pd.read_excel(file_path, sheet_name=None)

        # Determine compounds sheet and validate
        if len(excel_data) == 1:
            # Single sheet file
            compounds_sheet = list(excel_data.values())[0]
            compounds_input = compounds_sheet
        else:
            # Multi-sheet file
            if "Compounds" in excel_data:
                compounds_input = excel_data
            else:
                QMessageBox.warning(
                    self,
                    "Warning",
                    "Multi-sheet Excel file found, but no 'Compounds' sheet detected. Using the first sheet for compounds.",
                )
                compounds_input = list(excel_data.values())[0]

        # Validate formula/SMILES agreement before loading
        compounds_df = compounds_input.get("Compounds", list(compounds_input.values())[0]) if isinstance(compounds_input, dict) else compounds_input
        col_lower = {col.lower(): col for col in compounds_df.columns}
        formula_col = next(
            (col_lower[c] for c in ("chemicalformula", "formula", "sum_formula", "molformula") if c in col_lower),
            None,
        )
        smiles_col = next(
            (col_lower[c] for c in ("smiles", "smi") if c in col_lower),
            None,
        )
        if formula_col and smiles_col:
            name_col = next(
                (col_lower[c] for c in ("name", "compound", "compound_name") if c in col_lower),
                "",
            )
            problematic = validate_formula_smiles_agreement(compounds_df, formula_col, smiles_col, name_col)
            if problematic:
                names_text = "\n".join(f"  \u2022 {n}" for n in problematic)
                result = QMessageBox.question(
                    self,
                    "Formula/SMILES Mismatch",
                    f"{len(problematic)} compound(s) have a mismatch between sum formula and SMILES:\n\n{names_text}\n\nDo you want to continue the import anyway?",
                    QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                    QMessageBox.StandardButton.No,
                )
                if result != QMessageBox.StandardButton.Yes:
                    return

        # Load compounds
        self.compound_manager.load_compounds(compounds_input)
        self.update_compounds_table()

        compound_count = len(self.compound_manager.get_compounds_data())
        self.statusBar().showMessage(f"Compounds loaded via drag & drop. Total: {compound_count} compounds")

        # Start monitoring the compound file for changes
        self.start_compound_file_monitoring(file_path)

    def load_stylesheet(self):
        """Load the CSS stylesheet and apply it to the entire application."""
        stylesheet_path = os.path.join(os.path.dirname(__file__), "style.css")
        if os.path.exists(stylesheet_path):
            with open(stylesheet_path, "r") as f:
                QApplication.instance().setStyleSheet(f.read())

    def closeEvent(self, event):
        """Clean up when closing the application"""
        # Stop monitoring compound file
        self.stop_compound_file_monitoring()

        # Close all EIC windows
        for window in self.eic_windows:
            window.close()
        event.accept()

    @staticmethod
    def _get_default_eic_settings():
        """Return the hardcoded default values for EIC/MSMS settings."""
        return {
            "mz_tolerance_ppm": 5.0,
            "separate_groups": False,
            "rt_shift_min": 1.0,
            "crop_rt_window": True,
            "normalize_samples": False,
            "legend_position": "Right",
            "eic_method": "Sum of all signals",
            "msms_filter_regex": "(FTMS|ITMS).*",
            "msms_filter_replacement": "\\1",
            "msms_similarity_method": "CosineHungarian",
            "msms_similarity_default_tolerance": 0.1,
            "msms_similarity_group_tolerances": [
                {"filter_type": "ITMS", "mz_tolerance": 0.5},
                {"filter_type": "FTMS", "mz_tolerance": 0.01},
            ],
            "show_msms_closest": False,
            "show_msms_3s": False,
            "show_msms_6s": False,
            "show_msms_9s": False,
            "show_msms_most_abundant_3s": True,
            "show_msms_most_abundant_6s": False,
            "show_msms_most_abundant_9s": False,
            "settings_templates": [],
        }

    @staticmethod
    def _get_default_memory_settings():
        """Return the hardcoded default values for memory settings."""
        return {
            "keep_in_memory": True,
            "parallel_tasks": 4,
        }

    def load_eic_defaults(self):
        """Load EIC window default settings"""
        _d = self._get_default_eic_settings()
        self.eic_defaults = {
            "mz_tolerance_ppm": float(self.settings.value("eic/mz_tolerance_ppm", _d["mz_tolerance_ppm"])),
            "separate_groups": self.settings.value("eic/separate_groups", _d["separate_groups"], type=bool),
            "rt_shift_min": float(self.settings.value("eic/rt_shift_min", _d["rt_shift_min"])),
            "crop_rt_window": self.settings.value("eic/crop_rt_window", _d["crop_rt_window"], type=bool),
            "normalize_samples": self.settings.value("eic/normalize_samples", _d["normalize_samples"], type=bool),
            "legend_position": self.settings.value("eic/legend_position", _d["legend_position"]),
            "eic_method": self.settings.value("eic/eic_method", _d["eic_method"]),
            "msms_filter_regex": self.settings.value("msms_filter/regex", _d["msms_filter_regex"]),
            "msms_filter_replacement": self.settings.value("msms_filter/replacement", _d["msms_filter_replacement"]),
            "msms_similarity_method": self.settings.value("msms_similarity/method", _d["msms_similarity_method"]),
            "msms_similarity_default_tolerance": float(
                self.settings.value(
                    "msms_similarity/default_tolerance",
                    _d["msms_similarity_default_tolerance"],
                )
            ),
            "msms_similarity_group_tolerances": json.loads(
                self.settings.value(
                    "msms_similarity/group_tolerances",
                    json.dumps(_d["msms_similarity_group_tolerances"]),
                )
            ),
            "show_msms_closest": self.settings.value("eic/show_msms_closest", _d["show_msms_closest"], type=bool),
            "show_msms_3s": self.settings.value("eic/show_msms_3s", _d["show_msms_3s"], type=bool),
            "show_msms_6s": self.settings.value("eic/show_msms_6s", _d["show_msms_6s"], type=bool),
            "show_msms_9s": self.settings.value("eic/show_msms_9s", _d["show_msms_9s"], type=bool),
            "show_msms_most_abundant_3s": self.settings.value("eic/show_msms_most_abundant_3s", _d["show_msms_most_abundant_3s"], type=bool),
            "show_msms_most_abundant_6s": self.settings.value("eic/show_msms_most_abundant_6s", _d["show_msms_most_abundant_6s"], type=bool),
            "show_msms_most_abundant_9s": self.settings.value("eic/show_msms_most_abundant_9s", _d["show_msms_most_abundant_9s"], type=bool),
            "settings_templates": json.loads(self.settings.value("eic/settings_templates", json.dumps(_d["settings_templates"]))),
        }

        # Load memory settings
        _m = self._get_default_memory_settings()
        self.memory_settings = {
            "keep_in_memory": self.settings.value("memory/keep_in_memory", _m["keep_in_memory"], type=bool),
            "parallel_tasks": int(self.settings.value("memory/parallel_tasks", _m["parallel_tasks"])),
        }

        # Apply memory settings to file manager
        self.file_manager.set_memory_mode(self.memory_settings["keep_in_memory"], auto_load=True)

    def save_eic_defaults(self):
        """Save EIC window default settings"""
        self.settings.setValue("eic/mz_tolerance_ppm", self.eic_defaults["mz_tolerance_ppm"])
        self.settings.setValue("eic/separate_groups", self.eic_defaults["separate_groups"])
        self.settings.setValue("eic/rt_shift_min", self.eic_defaults["rt_shift_min"])
        self.settings.setValue("eic/crop_rt_window", self.eic_defaults["crop_rt_window"])
        self.settings.setValue("eic/normalize_samples", self.eic_defaults["normalize_samples"])
        self.settings.setValue("eic/legend_position", self.eic_defaults.get("legend_position", "Right"))
        self.settings.setValue("eic/eic_method", self.eic_defaults.get("eic_method", "Sum of all signals"))
        self.settings.setValue("msms_filter/regex", self.eic_defaults.get("msms_filter_regex", ""))
        self.settings.setValue(
            "msms_filter/replacement",
            self.eic_defaults.get("msms_filter_replacement", ""),
        )
        self.settings.setValue(
            "msms_similarity/method",
            self.eic_defaults.get("msms_similarity_method", "CosineHungarian"),
        )
        self.settings.setValue(
            "msms_similarity/default_tolerance",
            self.eic_defaults.get("msms_similarity_default_tolerance", 0.1),
        )
        self.settings.setValue(
            "msms_similarity/group_tolerances",
            json.dumps(self.eic_defaults.get("msms_similarity_group_tolerances", [])),
        )
        self.settings.setValue("eic/show_msms_closest", self.eic_defaults.get("show_msms_closest", True))
        self.settings.setValue("eic/show_msms_3s", self.eic_defaults.get("show_msms_3s", False))
        self.settings.setValue("eic/show_msms_6s", self.eic_defaults.get("show_msms_6s", False))
        self.settings.setValue("eic/show_msms_9s", self.eic_defaults.get("show_msms_9s", False))
        self.settings.setValue("eic/show_msms_most_abundant_3s", self.eic_defaults.get("show_msms_most_abundant_3s", False))
        self.settings.setValue("eic/show_msms_most_abundant_6s", self.eic_defaults.get("show_msms_most_abundant_6s", False))
        self.settings.setValue("eic/show_msms_most_abundant_9s", self.eic_defaults.get("show_msms_most_abundant_9s", False))
        self.settings.setValue("eic/settings_templates", json.dumps(self.eic_defaults.get("settings_templates", [])))

    def _on_eic_settings_changed(self, key: str, value) -> None:
        """Called by an EIC window when the user changes a persistent setting."""
        self.eic_defaults[key] = value
        self.save_eic_defaults()

    def save_memory_settings(self):
        """Save memory settings"""
        self.settings.setValue("memory/keep_in_memory", self.memory_settings["keep_in_memory"])
        self.settings.setValue("memory/parallel_tasks", self.memory_settings["parallel_tasks"])

    def update_memory_label(self):
        """Update the memory usage label"""
        try:
            memory_info = self.file_manager.get_memory_usage()

            if "error" in memory_info:
                self.memory_label.setText("Memory: Error")
            else:
                rss_mb = memory_info["rss_mb"]
                cached_files = memory_info["cached_files"]

                if memory_info["keep_in_memory"] and cached_files > 0:
                    self.memory_label.setText(f"Memory: {rss_mb:.1f} MB ({cached_files} files cached)")
                else:
                    self.memory_label.setText(f"Memory: {rss_mb:.1f} MB")
        except Exception as e:
            self.memory_label.setText("Memory: Error")
        self.settings.sync()

    def start_compound_file_monitoring(self, file_path):
        """Start monitoring the compound file for changes"""
        import os

        # Stop any existing monitoring
        self.stop_compound_file_monitoring()

        # Store file information
        self.compound_file_path = file_path
        try:
            stat = os.stat(file_path)
            self.compound_file_size = stat.st_size
            self.compound_file_mtime = stat.st_mtime
            self.compound_file_monitoring_active = True

            # Create and start timer (check every 2 seconds)
            self.compound_file_monitor_timer = QTimer()
            self.compound_file_monitor_timer.timeout.connect(self.check_compound_file_changes)
            self.compound_file_monitor_timer.start(2000)  # 2000 ms = 2 seconds

        except Exception as e:
            # If we can't stat the file, don't start monitoring
            self.compound_file_path = None
            self.compound_file_size = None
            self.compound_file_mtime = None
            self.compound_file_monitoring_active = False

    def stop_compound_file_monitoring(self):
        """Stop monitoring the compound file"""
        if self.compound_file_monitor_timer:
            self.compound_file_monitor_timer.stop()
            self.compound_file_monitor_timer = None

        self.compound_file_path = None
        self.compound_file_size = None
        self.compound_file_mtime = None
        self.compound_file_monitoring_active = False

    def check_compound_file_changes(self):
        """Check if the compound file has changed"""
        import os

        if not self.compound_file_monitoring_active or not self.compound_file_path:
            return

        try:
            # Check if file still exists
            if not os.path.exists(self.compound_file_path):
                return

            # Get current file stats
            stat = os.stat(self.compound_file_path)
            current_size = stat.st_size
            current_mtime = stat.st_mtime

            # Check if file has changed (size or modification time)
            if current_size != self.compound_file_size or current_mtime != self.compound_file_mtime:
                # File has changed - prompt user
                reply = QMessageBox.question(
                    self,
                    "Compound File Changed",
                    f"The compound file has been modified:\n\n{self.compound_file_path}\n\nDo you want to re-import it?",
                    QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                    QMessageBox.StandardButton.Yes,
                )

                if reply == QMessageBox.StandardButton.Yes:
                    # Clear existing compounds
                    self.compound_manager.compounds_data = pd.DataFrame()
                    self.compound_manager.compound_adduct_data = {}

                    # Re-import the file
                    self.load_compounds_from_file(self.compound_file_path)

                    # Update file stats after reload
                    stat = os.stat(self.compound_file_path)
                    self.compound_file_size = stat.st_size
                    self.compound_file_mtime = stat.st_mtime
                else:
                    # User declined - stop monitoring
                    self.stop_compound_file_monitoring()
                    self.statusBar().showMessage("Compound file monitoring stopped.")

        except Exception as e:
            # Error accessing file - silently continue monitoring
            pass

    def show_options_dialog(self):
        """Show the unified options configuration dialog"""
        dialog = UnifiedOptionsDialog(self.eic_defaults, self.memory_settings, self.file_manager, self)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            new_eic, new_mem = dialog.get_values()
            self.eic_defaults.update(new_eic)
            self.memory_settings.update(new_mem)
            self.save_eic_defaults()
            self.save_memory_settings()

            # Apply memory settings to file manager
            old_memory_mode = self.file_manager.keep_in_memory
            new_memory_mode = self.memory_settings["keep_in_memory"]

            if old_memory_mode != new_memory_mode:
                if new_memory_mode:
                    # Show loading dialog when enabling memory mode
                    self.file_manager.set_memory_mode(True, auto_load=False)  # Enable mode without auto-loading
                    self.load_files_to_memory_with_progress()
                else:
                    # Just disable memory mode
                    self.file_manager.set_memory_mode(False)

            QMessageBox.information(
                self,
                "Settings Saved",
                "Options have been saved and applied successfully!",
            )

    def _load_single_file_worker(self, row_data):
        """Worker function to load a single file in a separate thread"""
        idx, row = row_data
        filepath = row["Filepath"]
        try:
            data = self.file_manager.load_single_file(filepath)
            return idx, filepath, data["ms1"], data["ms2"]
        except Exception as e:
            print(f"Error loading file {filepath}: {e}")
            return idx, filepath, [], []

    def load_files_to_memory_with_progress(self):
        """Load files to memory with a progress dialog using parallel processing"""
        if self.file_manager.files_data.empty:
            return

        num_files = len(self.file_manager.files_data)
        parallel_tasks = self.memory_settings.get("parallel_tasks", 4)

        # Create progress dialog
        progress = QProgressDialog("Loading mzML files into memory...", "Cancel", 0, num_files, self)
        progress.setWindowTitle("Loading Files")
        progress.setWindowModality(Qt.WindowModality.WindowModal)
        progress.setMinimumDuration(0)
        progress.show()

        # Clear any existing cache first
        self.file_manager._clear_memory_cache()

        # Prepare data for parallel execution
        rows = list(self.file_manager.files_data.iterrows())

        with concurrent.futures.ThreadPoolExecutor(max_workers=parallel_tasks) as executor:
            futures = {executor.submit(self._load_single_file_worker, row): row for row in rows}

            completed_count = 0
            for future in concurrent.futures.as_completed(futures):
                if progress.wasCanceled():
                    executor.shutdown(wait=False)
                    self.file_manager.set_memory_mode(False)
                    QMessageBox.information(self, "Cancelled", "Memory loading was cancelled.")
                    return

                try:
                    idx, filepath, ms1_data, ms2_data = future.result()

                    # Store in cache
                    self.file_manager.cached_data[filepath] = {
                        "ms1": ms1_data,
                        "ms2": ms2_data,
                    }

                    completed_count += 1
                    progress.setValue(completed_count)
                    filename = os.path.basename(filepath)
                    progress.setLabelText(f"Loaded {filename} ({completed_count}/{num_files})")
                    QApplication.processEvents()

                except Exception as e:
                    print(f"Error in future result: {e}")

        progress.setValue(num_files)
        progress.close()

    def record_peak_integration(
        self,
        compound_name,
        ion_name,
        mz_value,
        rt_value,
        ion_mode,
        sample_name,
        group_name,
        peak_start_rt,
        peak_end_rt,
        peak_area,
    ):
        """
        Record peak integration data for a compound and ion.
        Only keeps the latest integration for each compound-ion combination.

        Args:
            compound_name: Name of the compound
            ion_name: Name/description of the ion (e.g., adduct type)
            mz_value: m/z value
            rt_value: retention time value (center)
            ion_mode: ionization mode (positive/negative)
            sample_name: name of the sample file
            group_name: group name for the sample
            peak_start_rt: start of peak boundary
            peak_end_rt: end of peak boundary
            peak_area: integrated peak area
        """
        key = (compound_name, ion_name)

        if key not in self.peak_integration_data:
            self.peak_integration_data[key] = {
                "compound_name": compound_name,
                "ion_name": ion_name,
                "mz_value": mz_value,
                "rt_value": rt_value,
                "ion_mode": ion_mode,
                "peak_start_rt": peak_start_rt,
                "peak_end_rt": peak_end_rt,
                "sample_data": [],  # List of (sample_name, group_name, peak_area) tuples
            }

        # Update the integration data - only keep latest boundaries and sample data
        integration_data = self.peak_integration_data[key]
        integration_data["peak_start_rt"] = peak_start_rt
        integration_data["peak_end_rt"] = peak_end_rt
        integration_data["sample_data"] = [(sample_name, group_name, peak_area)]

    def update_peak_integration_samples(self, compound_name, ion_name, sample_data_list):
        """
        Update the sample data for a compound-ion integration.

        Args:
            compound_name: Name of the compound
            ion_name: Name/description of the ion
            sample_data_list: List of (sample_name, group_name, peak_area) tuples
        """
        key = (compound_name, ion_name)

        if key in self.peak_integration_data:
            self.peak_integration_data[key]["sample_data"] = sample_data_list

    def export_peak_integration_excel(self):
        """Export peak integration data to Excel file in long format"""
        if not self.peak_integration_data:
            QMessageBox.information(self, "No Data", "No peak integration data available to export.")
            return

        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Export Peak Integration Data",
            "peak_integration_data.xlsx",
            "Excel files (*.xlsx)",
        )

        if not file_path:
            return

        try:
            # Prepare data for long format table
            export_data = []

            for (
                compound_name,
                ion_name,
            ), integration_data in self.peak_integration_data.items():
                base_row = {
                    "compound_name": integration_data["compound_name"],
                    "ion_name": integration_data["ion_name"],
                    "mz_value": integration_data["mz_value"],
                    "rt_value": integration_data["rt_value"],
                    "ion_mode": integration_data["ion_mode"],
                    "peak_start_rt": integration_data["peak_start_rt"],
                    "peak_end_rt": integration_data["peak_end_rt"],
                }

                # Add a row for each sample
                for sample_name, group_name, peak_area in integration_data["sample_data"]:
                    row = base_row.copy()
                    row.update(
                        {
                            "sample_name": sample_name,
                            "group_name": group_name,
                            "peak_area": peak_area,
                        }
                    )
                    export_data.append(row)

            # Create DataFrame and export
            df = pd.DataFrame(export_data)
            df.to_excel(file_path, index=False)

            QMessageBox.information(
                self,
                "Export Complete",
                f"Peak integration data exported to:\n{file_path}",
            )

        except Exception as e:
            QMessageBox.critical(self, "Export Error", f"Failed to export data:\n{str(e)}")

    def generate_r_code(self):
        """Generate R code for loading the exported Excel file"""
        if not self.peak_integration_data:
            QMessageBox.information(
                self,
                "No Data",
                "No peak integration data available to generate R code for.",
            )
            return

        r_code = """# R code to load peak integration data
# Install readxl package if not already installed
# install.packages("readxl")

library(readxl)
library(dplyr)

# Load the peak integration data
peak_data <- read_excel("peak_integration_data.xlsx")

# Display structure of the data
str(peak_data)

# Summary of the data
summary(peak_data)

# View first few rows
head(peak_data)

# Example: Group by compound and ion to get summary statistics
peak_summary <- peak_data %>%
  group_by(compound_name, ion_name, group_name) %>%
  summarise(
    n_samples = n(),
    mean_area = mean(peak_area, na.rm = TRUE),
    sd_area = sd(peak_area, na.rm = TRUE),
    median_area = median(peak_area, na.rm = TRUE),
    .groups = 'drop'
  )

print(peak_summary)

# Example: Create boxplot by groups
library(ggplot2)

ggplot(peak_data, aes(x = group_name, y = peak_area)) +
  geom_boxplot() +
  geom_jitter() +
  facet_wrap(~ paste(compound_name, ion_name, sep = " - "), scales = "free_y") +
  theme_minimal() +
  labs(
    title = "Peak Areas by Group",
    x = "Group",
    y = "Peak Area",
    caption = "Data from mzML Explorer peak integration"
  ) +
  theme(axis.text.x = element_text(angle = 45, hjust = 1))
"""

        # Show R code in a dialog
        dialog = QDialog(self)
        dialog.setWindowTitle("R Code for Loading Peak Integration Data")
        dialog.resize(800, 600)

        layout = QVBoxLayout(dialog)

        # Text area with R code
        from PyQt6.QtWidgets import QTextEdit

        text_edit = QTextEdit()
        text_edit.setPlainText(r_code)
        text_edit.setFont(QFont("Courier", 10))
        layout.addWidget(text_edit)

        # Buttons
        button_layout = QHBoxLayout()

        copy_button = QPushButton("Copy to Clipboard")
        copy_button.clicked.connect(lambda: QApplication.clipboard().setText(r_code))
        button_layout.addWidget(copy_button)

        save_button = QPushButton("Save to File")
        save_button.clicked.connect(lambda: self._save_r_code(r_code))
        button_layout.addWidget(save_button)

        button_layout.addStretch()

        close_button = QPushButton("Close")
        close_button.clicked.connect(dialog.close)
        button_layout.addWidget(close_button)

        layout.addLayout(button_layout)

        dialog.exec()

    # ------------------------------------------------------------------
    # Fetch compound information from online databases
    # ------------------------------------------------------------------

    def fetch_compound_information(self):
        """Fetch compound information from PubChem and enrich an Excel compounds file."""
        from .compound_info_fetcher import fetch_pubchem_info_batch, load_cas_api_key

        # Step 1: Select input Excel file
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "Select Compounds Excel File to Enrich",
            "",
            "Excel files (*.xlsx);;All files (*)",
        )
        if not file_path:
            return

        # Step 2: Read the Excel file
        try:
            df = pd.read_excel(file_path)
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to read Excel file:\n{str(e)}")
            return

        if df.empty:
            QMessageBox.information(self, "Empty File", "The selected Excel file is empty.")
            return

        # Step 3: Detect relevant columns (case-insensitive)
        col_lower = {col.lower(): col for col in df.columns}

        name_col = col_lower.get("name")
        if not name_col:
            QMessageBox.warning(
                self,
                "Missing Column",
                "No 'Name' column found in the selected Excel file.",
            )
            return

        # Detect optional CAS column
        cas_col = next(
            (col_lower[k] for k in ("cas", "cas_number", "cas_no", "cas number", "casno") if k in col_lower),
            None,
        )

        # Detect existing SMILES / formula / mass columns
        smiles_col = col_lower.get("smiles") or col_lower.get("canonical_smiles")
        formula_col = col_lower.get("chemicalformula") or col_lower.get("chemical_formula") or col_lower.get("formula")
        # Step 4: Collect (name, cas) pairs, then batch-fetch from PubChem
        n = len(df)
        compounds_to_fetch = []
        for _, row in df.iterrows():
            name = str(row[name_col]).strip() if pd.notna(row[name_col]) else ""
            cas = None
            if cas_col and pd.notna(row.get(cas_col)):
                cas_raw = str(row[cas_col]).strip()
                cas = cas_raw if cas_raw not in ("", "nan", "None", "NaN") else None
            compounds_to_fetch.append((name or None, cas))

        progress = QProgressDialog(
            f"Resolving compound identifiers (0/{n})...",
            "Cancel",
            0,
            n,
            self,
        )
        progress.setWindowTitle("Fetching Compound Information")
        progress.setWindowModality(Qt.WindowModality.WindowModal)
        progress.setMinimumDuration(0)
        progress.setValue(0)

        def _on_cid_resolved(i: int, total: int, name: str) -> bool:
            if progress.wasCanceled():
                return False
            progress.setLabelText(f"Resolving CID: {name or '(unnamed)'} ({i + 1}/{total})...")
            progress.setValue(i)
            QApplication.processEvents()
            return True

        # CID resolution is sequential (one lookup per compound), but property
        # and synonym data are fetched in a single batch request each, greatly
        # reducing the total number of HTTP calls for large compound lists.
        # CAS enrichment is performed afterwards when an API key is available.
        cas_api_key = load_cas_api_key()
        fetched_results = fetch_pubchem_info_batch(
            compounds_to_fetch,
            progress_callback=_on_cid_resolved,
            cas_api_key=cas_api_key,
        )
        cancelled = progress.wasCanceled()

        progress.setValue(n)
        progress.close()

        if cancelled:
            reply = QMessageBox.question(
                self,
                "Cancelled",
                "Fetching was cancelled. Apply partial results to the file?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            )
            if reply != QMessageBox.StandardButton.Yes:
                return

        # --- Helper: canonicalize a SMILES string via rdkit (if available) ---
        def _canonical_smiles(smi: str):
            try:
                from rdkit import Chem

                mol = Chem.MolFromSmiles(smi)
                if mol:
                    return Chem.MolToSmiles(mol)
            except Exception:
                pass
            return None

        # Step 5: Map fetched fields to target columns
        # Each entry: (fetched_key, target_column_in_df)
        column_mapping = [
            ("cid", "PubChem_CID"),
            ("smiles", smiles_col or "SMILES"),
            ("molecular_formula", formula_col or "ChemicalFormula"),
            ("iupac_name", "IUPAC_Name"),
            ("inchi", "InChI"),
            ("inchikey", "InChIKey"),
            ("xlogp", "XLogP"),
            ("charge", "Charge"),
            ("literature_count", "LiteratureCount"),
            ("fingerprint2d", "Fingerprint2D"),
            ("title", "Title"),
            ("synonyms", "Synonyms"),
            ("cas_number", cas_col or "CAS"),
            ("cas_preferred_name", "CAS_Preferred_Name"),
            ("cas_rn_confirmed", "CAS_RN_Confirmed"),
            ("cas_experimental_properties", "CAS_Experimental_Properties"),
        ]

        # Ensure all target columns exist in the DataFrame
        for _, target_col in column_mapping:
            if target_col not in df.columns:
                df[target_col] = None

        cell_colors: dict = {}  # (df_row_position, col_name) -> "yellow" | "orange"

        for row_pos, (df_idx, row) in enumerate(df.iterrows()):
            info = fetched_results[row_pos] if row_pos < len(fetched_results) else None
            if info is None:
                continue

            for fetched_key, target_col in column_mapping:
                fetched_val = info.get(fetched_key)
                if fetched_val is None:
                    continue

                # Serialize lists (synonyms) as a JSON array
                if isinstance(fetched_val, list):
                    fetched_str = json.dumps(fetched_val, ensure_ascii=False)
                elif isinstance(fetched_val, (int, float)):
                    fetched_str = fetched_val
                else:
                    fetched_str = str(fetched_val).strip()

                existing_val = row[target_col]
                is_empty = existing_val is None or (isinstance(existing_val, float) and pd.isna(existing_val)) or str(existing_val).strip() in ("", "nan", "None", "NaN")

                if is_empty:
                    df.at[df_idx, target_col] = fetched_str
                    cell_colors[(row_pos, target_col)] = "yellow"
                else:
                    existing_str = str(existing_val).strip()
                    fetched_compare = str(fetched_str).strip()

                    # For SMILES columns, compare canonical forms via rdkit
                    if fetched_key == "smiles":
                        existing_canon = _canonical_smiles(existing_str)
                        fetched_canon = _canonical_smiles(fetched_compare)
                        if existing_canon and fetched_canon and existing_canon == fetched_canon:
                            # Same molecule – update to PubChem canonical form if text differs
                            if existing_str != fetched_compare:
                                df.at[df_idx, target_col] = fetched_str
                                cell_colors[(row_pos, target_col)] = "yellow"
                            continue  # not a conflict

                    if existing_str != fetched_compare:
                        df.at[df_idx, target_col] = f"available: {existing_str} $$$ fetched: {fetched_compare}"
                        cell_colors[(row_pos, target_col)] = "orange"
                    # exact match: no change, no colour

        # Step 6: Ask user where to save the enriched file
        default_save = file_path.rsplit(".", 1)[0] + "_enriched.xlsx"
        save_path, _ = QFileDialog.getSaveFileName(
            self,
            "Save Enriched Compounds File",
            default_save,
            "Excel files (*.xlsx);;All files (*)",
        )
        if not save_path:
            return

        # Step 7: Write Excel with colour-coded cells
        try:
            self._save_enriched_excel(df, save_path, cell_colors)
            n_yellow = sum(1 for v in cell_colors.values() if v == "yellow")
            n_orange = sum(1 for v in cell_colors.values() if v == "orange")
            QMessageBox.information(
                self,
                "Done",
                (
                    f"Enriched compound file saved to:\n{save_path}\n\n  \u2022 {n_yellow} cell(s) with new information (yellow)\n  \u2022 {n_orange} cell(s) with conflicting information (orange)"
                ),
            )
        except Exception as e:
            QMessageBox.critical(self, "Save Error", f"Failed to save file:\n{str(e)}")

    def _save_enriched_excel(self, df: pd.DataFrame, path: str, cell_colors: dict):
        """Write *df* to an Excel file at *path*, applying yellow/orange cell fills."""
        import openpyxl
        from openpyxl.styles import PatternFill

        YELLOW = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        ORANGE = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Compounds"

        # Header row
        for col_idx, col_name in enumerate(df.columns, start=1):
            ws.cell(row=1, column=col_idx, value=col_name)

        col_index = {col: idx + 1 for idx, col in enumerate(df.columns)}

        # Data rows  (Excel row 2 == DataFrame position 0)
        for row_pos, (_, row) in enumerate(df.iterrows()):
            excel_row = row_pos + 2
            for col_name in df.columns:
                val = row[col_name]
                # Convert NaN / NaT to None for a clean spreadsheet
                if isinstance(val, float) and pd.isna(val):
                    val = None
                cell = ws.cell(row=excel_row, column=col_index[col_name], value=val)
                color = cell_colors.get((row_pos, col_name))
                if color == "yellow":
                    cell.fill = YELLOW
                elif color == "orange":
                    cell.fill = ORANGE

        wb.save(path)

    def _save_r_code(self, r_code):
        """Save R code to a file"""
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Save R Code",
            "peak_integration_analysis.R",
            "R files (*.R);;Text files (*.txt);;All files (*)",
        )

        if file_path:
            try:
                with open(file_path, "w") as f:
                    f.write(r_code)
                QMessageBox.information(self, "Save Complete", f"R code saved to:\n{file_path}")
            except Exception as e:
                QMessageBox.critical(self, "Save Error", f"Failed to save R code:\n{str(e)}")


class UnifiedOptionsDialog(QDialog):
    """Unified dialog for all application options with collapsible sections"""

    def __init__(self, eic_defaults, memory_settings, file_manager, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Options")
        self.setModal(True)
        self.setMinimumSize(600, 500)

        self.eic_defaults = eic_defaults.copy()
        self.memory_settings = memory_settings.copy()
        self.file_manager = file_manager
        self.init_ui()

    def init_ui(self):
        """Initialize the dialog UI"""
        layout = QVBoxLayout(self)

        # Create scroll area for the content
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)

        content_widget = QWidget()
        content_layout = QVBoxLayout(content_widget)

        # Memory Settings Section
        memory_section = CollapsibleBox("Memory Settings")
        memory_content = self.create_memory_settings_content()
        memory_section.add_widget(memory_content)
        memory_section.set_expanded(True)
        content_layout.addWidget(memory_section)

        # EIC Defaults Section
        eic_section = CollapsibleBox("EIC Window Defaults")
        eic_content = self.create_eic_defaults_content()
        eic_section.add_widget(eic_content)
        eic_section.set_expanded(True)
        content_layout.addWidget(eic_section)

        # MSMS Filter String Parsing Section
        msms_filter_section = CollapsibleBox("MSMS Filter String Parsing")
        msms_filter_content = self.create_msms_filter_content()
        msms_filter_section.add_widget(msms_filter_content)
        msms_filter_section.set_expanded(True)
        content_layout.addWidget(msms_filter_section)

        # MSMS Similarity Scoring Section
        msms_similarity_section = CollapsibleBox("MSMS Similarity Scoring")
        msms_similarity_content = self.create_msms_similarity_content()
        msms_similarity_section.add_widget(msms_similarity_content)
        msms_similarity_section.set_expanded(True)
        content_layout.addWidget(msms_similarity_section)

        content_layout.addStretch()
        scroll.setWidget(content_widget)
        layout.addWidget(scroll)

        # Buttons
        button_layout = QHBoxLayout()

        self.reset_button = QPushButton("Reset All to Defaults")
        self.reset_button.clicked.connect(self.reset_all_to_defaults)

        self.ok_button = QPushButton("OK")
        self.ok_button.clicked.connect(self.accept)

        self.cancel_button = QPushButton("Cancel")
        self.cancel_button.clicked.connect(self.reject)

        button_layout.addWidget(self.reset_button)
        button_layout.addStretch()
        button_layout.addWidget(self.cancel_button)
        button_layout.addWidget(self.ok_button)

        layout.addLayout(button_layout)

    def create_memory_settings_content(self):
        """Create the memory settings content widget"""
        widget = QWidget()
        layout = QVBoxLayout(widget)

        # Description
        desc_label = QLabel("Configure how mzML data is handled in memory. Keeping data in memory provides faster access but uses more RAM.")
        desc_label.setWordWrap(True)
        desc_label.setStyleSheet("QLabel { color: #555; margin-bottom: 10px; }")
        layout.addWidget(desc_label)

        # Form layout
        form_layout = QFormLayout()

        # Keep in memory checkbox
        self.keep_in_memory_cb = QCheckBox()
        self.keep_in_memory_cb.setChecked(self.memory_settings["keep_in_memory"])
        self.keep_in_memory_cb.toggled.connect(self.on_memory_mode_changed)
        form_layout.addRow("Keep all mzML data in memory:", self.keep_in_memory_cb)

        # Parallel tasks spinbox
        self.parallel_tasks_spin = NoScrollSpinBox()
        self.parallel_tasks_spin.setRange(1, 32)
        self.parallel_tasks_spin.setValue(self.memory_settings.get("parallel_tasks", 4))
        self.parallel_tasks_spin.setSuffix(" threads")
        form_layout.addRow("Number of parallel tasks:", self.parallel_tasks_spin)

        layout.addLayout(form_layout)

        # Current memory usage
        memory_group = QGroupBox("Current Memory Usage")
        memory_layout = QVBoxLayout(memory_group)

        memory_info = self.file_manager.get_memory_usage()

        if "error" in memory_info:
            memory_text = f"Error getting memory info: {memory_info['error']}"
        else:
            rss_mb = memory_info["rss_mb"]
            cached_files = memory_info["cached_files"]
            memory_text = f"Current memory usage: {rss_mb:.1f} MB\n"
            memory_text += f"Cached files: {cached_files}\n"
            memory_text += f"Memory mode: {'Enabled' if memory_info['keep_in_memory'] else 'Disabled'}"

        self.memory_info_label = QLabel(memory_text)
        self.memory_info_label.setStyleSheet("QLabel { font-family: monospace; }")
        memory_layout.addWidget(self.memory_info_label)

        # Refresh button
        refresh_button = QPushButton("Refresh Memory Info")
        refresh_button.clicked.connect(self.refresh_memory_info)
        memory_layout.addWidget(refresh_button)

        layout.addWidget(memory_group)

        # Warning label
        self.memory_warning_label = QLabel()
        self.memory_warning_label.setWordWrap(True)
        self.memory_warning_label.setStyleSheet("QLabel { color: #d66; margin: 10px 0; }")
        self.update_memory_warning()
        layout.addWidget(self.memory_warning_label)

        return widget

    def create_eic_defaults_content(self):
        """Create the EIC defaults content widget"""
        widget = QWidget()
        layout = QVBoxLayout(widget)

        # Description
        desc_label = QLabel("Configure default settings for new EIC (Extracted Ion Chromatogram) windows.")
        desc_label.setWordWrap(True)
        desc_label.setStyleSheet("QLabel { color: #555; margin-bottom: 10px; }")
        layout.addWidget(desc_label)

        # Form layout
        form_layout = QFormLayout()

        # m/z Tolerance (ppm)
        self.mz_tolerance_spin = NoScrollDoubleSpinBox()
        self.mz_tolerance_spin.setRange(0.1, 10000.0)
        self.mz_tolerance_spin.setValue(self.eic_defaults["mz_tolerance_ppm"])
        self.mz_tolerance_spin.setSuffix(" ppm")
        self.mz_tolerance_spin.setDecimals(1)
        self.mz_tolerance_spin.setSingleStep(1.0)
        form_layout.addRow("m/z Tolerance:", self.mz_tolerance_spin)

        # Separate by groups
        self.separate_groups_cb = QCheckBox()
        self.separate_groups_cb.setChecked(self.eic_defaults["separate_groups"])
        form_layout.addRow("Separate by groups:", self.separate_groups_cb)

        # Group RT Shift
        self.rt_shift_spin = NoScrollDoubleSpinBox()
        self.rt_shift_spin.setRange(0.0, 60.0)
        self.rt_shift_spin.setValue(self.eic_defaults["rt_shift_min"])
        self.rt_shift_spin.setSuffix(" min")
        self.rt_shift_spin.setDecimals(1)
        form_layout.addRow("Group RT Shift:", self.rt_shift_spin)

        # Crop to RT Window
        self.crop_rt_cb = QCheckBox()
        self.crop_rt_cb.setChecked(self.eic_defaults["crop_rt_window"])
        form_layout.addRow("Crop to RT Window:", self.crop_rt_cb)

        # Normalize to Max per Sample
        self.normalize_cb = QCheckBox()
        self.normalize_cb.setChecked(self.eic_defaults["normalize_samples"])
        form_layout.addRow("Normalize to Max per Sample:", self.normalize_cb)

        layout.addLayout(form_layout)

        # MSMS context-menu options
        msms_group = QGroupBox("MSMS Context Menu Options")
        msms_group_layout = QFormLayout(msms_group)

        self.show_msms_closest_cb = QCheckBox()
        self.show_msms_closest_cb.setChecked(self.eic_defaults.get("show_msms_closest", True))
        msms_group_layout.addRow("Show closest MSMS scan:", self.show_msms_closest_cb)

        self.show_msms_3s_cb = QCheckBox()
        self.show_msms_3s_cb.setChecked(self.eic_defaults.get("show_msms_3s", False))
        msms_group_layout.addRow("Show ±3 seconds around MSMS scan:", self.show_msms_3s_cb)

        self.show_msms_6s_cb = QCheckBox()
        self.show_msms_6s_cb.setChecked(self.eic_defaults.get("show_msms_6s", False))
        msms_group_layout.addRow("Show ±6 seconds around MSMS scan:", self.show_msms_6s_cb)

        self.show_msms_9s_cb = QCheckBox()
        self.show_msms_9s_cb.setChecked(self.eic_defaults.get("show_msms_9s", False))
        msms_group_layout.addRow("Show ±9 seconds around MSMS scan:", self.show_msms_9s_cb)

        msms_group_layout.addRow(QLabel("<hr>"))
        most_abundant_label = QLabel("<i>Select single most-abundant MSMS spectrum per file in RT window</i>")
        most_abundant_label.setWordWrap(True)
        msms_group_layout.addRow(most_abundant_label)

        self.show_msms_most_abundant_3s_cb = QCheckBox()
        self.show_msms_most_abundant_3s_cb.setChecked(self.eic_defaults.get("show_msms_most_abundant_3s", False))
        msms_group_layout.addRow("Most abundant MSMS per file (±3 s):", self.show_msms_most_abundant_3s_cb)

        self.show_msms_most_abundant_6s_cb = QCheckBox()
        self.show_msms_most_abundant_6s_cb.setChecked(self.eic_defaults.get("show_msms_most_abundant_6s", False))
        msms_group_layout.addRow("Most abundant MSMS per file (±6 s):", self.show_msms_most_abundant_6s_cb)

        self.show_msms_most_abundant_9s_cb = QCheckBox()
        self.show_msms_most_abundant_9s_cb.setChecked(self.eic_defaults.get("show_msms_most_abundant_9s", False))
        msms_group_layout.addRow("Most abundant MSMS per file (±9 s):", self.show_msms_most_abundant_9s_cb)

        layout.addWidget(msms_group)

        return widget

    def create_msms_filter_content(self):
        """Create the MSMS filter string parsing content widget"""
        widget = QWidget()
        layout = QVBoxLayout(widget)

        desc_label = QLabel(
            "Define a regex to parse the Orbitrap/instrument filter string of each MS2 spectrum "
            "into a short type label.\n"
            "Match regex: a Python regex applied to the filter string (may use capture groups).\n"
            "Type label: replacement string using \\1, \\2 … to form the type label from groups.\n"
            "Example — Match: r'@(\\w+?)\\d' → Label: \\1  produces 'hcd', 'cid', etc."
        )
        desc_label.setWordWrap(True)
        desc_label.setStyleSheet("QLabel { color: #555; margin-bottom: 10px; }")
        layout.addWidget(desc_label)

        form_layout = QFormLayout()

        self.msms_filter_regex_edit = QLineEdit()
        self.msms_filter_regex_edit.setPlaceholderText("e.g.  @(\\w+?)\\d")
        self.msms_filter_regex_edit.setText(self.eic_defaults.get("msms_filter_regex", ""))
        form_layout.addRow("Match regex:", self.msms_filter_regex_edit)

        self.msms_filter_replacement_edit = QLineEdit()
        self.msms_filter_replacement_edit.setPlaceholderText("e.g.  \\1")
        self.msms_filter_replacement_edit.setText(self.eic_defaults.get("msms_filter_replacement", ""))
        form_layout.addRow("Type label:", self.msms_filter_replacement_edit)

        layout.addLayout(form_layout)

        # Live test row
        test_layout = QHBoxLayout()
        self.msms_filter_test_input = QLineEdit()
        self.msms_filter_test_input.setPlaceholderText("Paste a filter string here to preview…")
        test_layout.addWidget(self.msms_filter_test_input)
        test_btn = QPushButton("Test")
        test_btn.setMaximumWidth(60)
        test_btn.clicked.connect(self._test_msms_filter_regex)
        test_layout.addWidget(test_btn)
        self.msms_filter_test_result = QLabel("")
        self.msms_filter_test_result.setStyleSheet("QLabel { color: #060; font-weight: bold; }")
        test_layout.addWidget(self.msms_filter_test_result)
        layout.addLayout(test_layout)

        return widget

    def create_msms_similarity_content(self):
        """Create the MSMS similarity scoring content widget."""
        widget = QWidget()
        layout = QVBoxLayout(widget)

        desc_label = QLabel(
            "Configure the cosine-similarity scoring method used when comparing MS2 spectra.\nYou can set a default m/z tolerance (Da) and optionally override it per filter-type group (requires the Match regex above to be configured)."
        )
        desc_label.setWordWrap(True)
        desc_label.setStyleSheet("QLabel { color: #555; margin-bottom: 8px; }")
        layout.addWidget(desc_label)

        form_layout = QFormLayout()

        self.msms_similarity_method_combo = NoScrollComboBox()
        self.msms_similarity_method_combo.addItems(["CosineHungarian", "CosineGreedy"])
        current_method = self.eic_defaults.get("msms_similarity_method", "CosineHungarian")
        idx = self.msms_similarity_method_combo.findText(current_method)
        if idx >= 0:
            self.msms_similarity_method_combo.setCurrentIndex(idx)
        form_layout.addRow("Scoring method:", self.msms_similarity_method_combo)

        self.msms_similarity_default_tol_spin = NoScrollDoubleSpinBox()
        self.msms_similarity_default_tol_spin.setRange(0.0001, 1.0)
        self.msms_similarity_default_tol_spin.setDecimals(4)
        self.msms_similarity_default_tol_spin.setSingleStep(0.005)
        self.msms_similarity_default_tol_spin.setSuffix(" Da")
        self.msms_similarity_default_tol_spin.setValue(self.eic_defaults.get("msms_similarity_default_tolerance", 0.1))
        form_layout.addRow("Default m/z tolerance:", self.msms_similarity_default_tol_spin)
        layout.addLayout(form_layout)

        # Per-group tolerance table
        group_tol_label = QLabel("<b>Per-group m/z tolerances</b> (optional — override per filter-type label):")
        group_tol_label.setWordWrap(True)
        layout.addWidget(group_tol_label)

        self.msms_similarity_group_tol_table = QTableWidget(0, 2)
        self.msms_similarity_group_tol_table.setHorizontalHeaderLabels(["Filter Type", "Tolerance (Da)"])
        self.msms_similarity_group_tol_table.horizontalHeader().setStretchLastSection(True)
        self.msms_similarity_group_tol_table.setMaximumHeight(160)
        self.msms_similarity_group_tol_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)

        # Populate with stored values
        for entry in self.eic_defaults.get("msms_similarity_group_tolerances", []):
            self._msms_similarity_add_table_row(entry.get("filter_type", ""), entry.get("mz_tolerance", 0.1))

        layout.addWidget(self.msms_similarity_group_tol_table)

        btn_layout = QHBoxLayout()
        add_row_btn = QPushButton("Add Row")
        add_row_btn.setMaximumWidth(80)
        add_row_btn.clicked.connect(lambda: self._msms_similarity_add_table_row("", 0.1))
        remove_row_btn = QPushButton("Remove Row")
        remove_row_btn.setMaximumWidth(90)
        remove_row_btn.clicked.connect(self._msms_similarity_remove_selected_row)
        btn_layout.addWidget(add_row_btn)
        btn_layout.addWidget(remove_row_btn)
        btn_layout.addStretch()
        layout.addLayout(btn_layout)

        return widget

    def _msms_similarity_add_table_row(self, filter_type: str, tolerance: float):
        """Append one row to the per-group tolerance table."""
        table = self.msms_similarity_group_tol_table
        row = table.rowCount()
        table.insertRow(row)
        table.setItem(row, 0, QTableWidgetItem(str(filter_type)))
        tol_item = QTableWidgetItem(str(tolerance))
        tol_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
        table.setItem(row, 1, tol_item)

    def _msms_similarity_remove_selected_row(self):
        """Remove the currently selected row from the per-group tolerance table."""
        table = self.msms_similarity_group_tol_table
        selected = table.selectedItems()
        if selected:
            table.removeRow(table.row(selected[0]))

    def _test_msms_filter_regex(self):
        """Test the current regex/replacement against the sample input and show the result."""
        pattern = self.msms_filter_regex_edit.text().strip()
        replacement = self.msms_filter_replacement_edit.text()
        test_input = self.msms_filter_test_input.text().strip()
        if not pattern:
            self.msms_filter_test_result.setText("(no pattern)")
            self.msms_filter_test_result.setStyleSheet("QLabel { color: #888; }")
            return
        try:
            compiled = re.compile(pattern)
            m = compiled.search(test_input)
            if m:
                label = m.expand(replacement)
                self.msms_filter_test_result.setText(f"→ '{label}'")
                self.msms_filter_test_result.setStyleSheet("QLabel { color: #060; font-weight: bold; }")
            else:
                self.msms_filter_test_result.setText("(no match)")
                self.msms_filter_test_result.setStyleSheet("QLabel { color: #a60; }")
        except re.error as exc:
            self.msms_filter_test_result.setText(f"Regex error: {exc}")
            self.msms_filter_test_result.setStyleSheet("QLabel { color: #c00; }")

    def on_memory_mode_changed(self):
        """Handle memory mode checkbox change"""
        self.update_memory_warning()

    def update_memory_warning(self):
        """Update the memory warning label based on current settings"""
        if self.keep_in_memory_cb.isChecked():
            num_files = len(self.file_manager.get_files_data())
            if num_files > 0:
                self.memory_warning_label.setText(
                    f"⚠ Warning: Enabling memory mode will load all {num_files} mzML files into RAM. This may use significant memory and take time to load."
                )
            else:
                self.memory_warning_label.setText("ℹ Note: Memory mode will load all mzML files into RAM when files are added.")
        else:
            if self.file_manager.keep_in_memory:
                self.memory_warning_label.setText("ℹ Disabling memory mode will clear the current cache and revert to file-based reading.")
            else:
                self.memory_warning_label.setText("")

    def refresh_memory_info(self):
        """Refresh the memory information display"""
        memory_info = self.file_manager.get_memory_usage()

        if "error" in memory_info:
            memory_text = f"Error getting memory info: {memory_info['error']}"
        else:
            rss_mb = memory_info["rss_mb"]
            cached_files = memory_info["cached_files"]
            memory_text = f"Current memory usage: {rss_mb:.1f} MB\n"
            memory_text += f"Cached files: {cached_files}\n"
            memory_text += f"Memory mode: {'Enabled' if memory_info['keep_in_memory'] else 'Disabled'}"

        self.memory_info_label.setText(memory_text)

    def reset_all_to_defaults(self):
        """Reset all values to application defaults"""
        defaults = MzMLExplorerMainWindow._get_default_eic_settings()
        mem_defaults = MzMLExplorerMainWindow._get_default_memory_settings()

        # Reset EIC defaults
        self.mz_tolerance_spin.setValue(defaults["mz_tolerance_ppm"])
        self.separate_groups_cb.setChecked(defaults["separate_groups"])
        self.rt_shift_spin.setValue(defaults["rt_shift_min"])
        self.crop_rt_cb.setChecked(defaults["crop_rt_window"])
        self.normalize_cb.setChecked(defaults["normalize_samples"])

        # Reset memory settings
        self.keep_in_memory_cb.setChecked(mem_defaults["keep_in_memory"])
        self.parallel_tasks_spin.setValue(mem_defaults["parallel_tasks"])
        self.update_memory_warning()

        # Reset MSMS filter settings
        self.msms_filter_regex_edit.setText(defaults["msms_filter_regex"])
        self.msms_filter_replacement_edit.setText(defaults["msms_filter_replacement"])

        # Reset MSMS similarity settings
        self.msms_similarity_method_combo.setCurrentText(defaults["msms_similarity_method"])
        self.msms_similarity_default_tol_spin.setValue(defaults["msms_similarity_default_tolerance"])
        self.msms_similarity_group_tol_table.setRowCount(0)
        for entry in defaults["msms_similarity_group_tolerances"]:
            self._msms_similarity_add_table_row(entry["filter_type"], entry["mz_tolerance"])

        # Reset MSMS context menu options
        self.show_msms_closest_cb.setChecked(defaults["show_msms_closest"])
        self.show_msms_3s_cb.setChecked(defaults["show_msms_3s"])
        self.show_msms_6s_cb.setChecked(defaults["show_msms_6s"])
        self.show_msms_9s_cb.setChecked(defaults["show_msms_9s"])
        self.show_msms_most_abundant_3s_cb.setChecked(defaults.get("show_msms_most_abundant_3s", False))
        self.show_msms_most_abundant_6s_cb.setChecked(defaults.get("show_msms_most_abundant_6s", False))
        self.show_msms_most_abundant_9s_cb.setChecked(defaults.get("show_msms_most_abundant_9s", False))

    def get_values(self):
        """Get the current values from the dialog"""
        # Collect per-group tolerances from the table
        table = self.msms_similarity_group_tol_table
        group_tolerances = []
        for row in range(table.rowCount()):
            ft_item = table.item(row, 0)
            tol_item = table.item(row, 1)
            if ft_item and tol_item:
                ft = ft_item.text().strip()
                try:
                    tol = float(tol_item.text())
                except ValueError:
                    tol = 0.1
                if ft:
                    group_tolerances.append({"filter_type": ft, "mz_tolerance": tol})

        eic_values = {
            "mz_tolerance_ppm": self.mz_tolerance_spin.value(),
            "separate_groups": self.separate_groups_cb.isChecked(),
            "rt_shift_min": self.rt_shift_spin.value(),
            "crop_rt_window": self.crop_rt_cb.isChecked(),
            "normalize_samples": self.normalize_cb.isChecked(),
            "msms_filter_regex": self.msms_filter_regex_edit.text(),
            "msms_filter_replacement": self.msms_filter_replacement_edit.text(),
            "msms_similarity_method": self.msms_similarity_method_combo.currentText(),
            "msms_similarity_default_tolerance": self.msms_similarity_default_tol_spin.value(),
            "msms_similarity_group_tolerances": group_tolerances,
            "show_msms_closest": self.show_msms_closest_cb.isChecked(),
            "show_msms_3s": self.show_msms_3s_cb.isChecked(),
            "show_msms_6s": self.show_msms_6s_cb.isChecked(),
            "show_msms_9s": self.show_msms_9s_cb.isChecked(),
            "show_msms_most_abundant_3s": self.show_msms_most_abundant_3s_cb.isChecked(),
            "show_msms_most_abundant_6s": self.show_msms_most_abundant_6s_cb.isChecked(),
            "show_msms_most_abundant_9s": self.show_msms_most_abundant_9s_cb.isChecked(),
        }

        memory_values = {
            "keep_in_memory": self.keep_in_memory_cb.isChecked(),
            "parallel_tasks": self.parallel_tasks_spin.value(),
        }

        return eic_values, memory_values


class EICDefaultsDialog(QDialog):
    """Dialog for configuring EIC window defaults"""

    def __init__(self, current_defaults, parent=None):
        super().__init__(parent)
        self.setWindowTitle("EIC Window Defaults")
        self.setModal(True)
        self.setFixedSize(400, 300)

        self.current_defaults = current_defaults.copy()
        self.init_ui()

    def init_ui(self):
        """Initialize the dialog UI"""
        layout = QVBoxLayout(self)

        # Create form layout
        form_layout = QFormLayout()

        # m/z Tolerance (ppm)
        self.mz_tolerance_spin = NoScrollDoubleSpinBox()
        self.mz_tolerance_spin.setRange(0.1, 10000.0)
        self.mz_tolerance_spin.setValue(self.current_defaults["mz_tolerance_ppm"])
        self.mz_tolerance_spin.setSuffix(" ppm")
        self.mz_tolerance_spin.setDecimals(1)
        self.mz_tolerance_spin.setSingleStep(1.0)
        form_layout.addRow("m/z Tolerance:", self.mz_tolerance_spin)

        # Separate by groups
        self.separate_groups_cb = QCheckBox()
        self.separate_groups_cb.setChecked(self.current_defaults["separate_groups"])
        form_layout.addRow("Separate by groups:", self.separate_groups_cb)

        # Group RT Shift
        self.rt_shift_spin = NoScrollDoubleSpinBox()
        self.rt_shift_spin.setRange(0.0, 60.0)
        self.rt_shift_spin.setValue(self.current_defaults["rt_shift_min"])
        self.rt_shift_spin.setSuffix(" min")
        self.rt_shift_spin.setDecimals(1)
        form_layout.addRow("Group RT Shift:", self.rt_shift_spin)

        # Crop to RT Window
        self.crop_rt_cb = QCheckBox()
        self.crop_rt_cb.setChecked(self.current_defaults["crop_rt_window"])
        form_layout.addRow("Crop to RT Window:", self.crop_rt_cb)

        # Normalize to Max per Sample
        self.normalize_cb = QCheckBox()
        self.normalize_cb.setChecked(self.current_defaults["normalize_samples"])
        form_layout.addRow("Normalize to Max per Sample:", self.normalize_cb)

        layout.addLayout(form_layout)

        # Buttons
        button_layout = QHBoxLayout()

        self.ok_button = QPushButton("OK")
        self.ok_button.clicked.connect(self.accept)

        self.cancel_button = QPushButton("Cancel")
        self.cancel_button.clicked.connect(self.reject)

        self.reset_button = QPushButton("Reset to Defaults")
        self.reset_button.clicked.connect(self.reset_to_defaults)

        button_layout.addWidget(self.reset_button)
        button_layout.addStretch()
        button_layout.addWidget(self.cancel_button)
        button_layout.addWidget(self.ok_button)

        layout.addLayout(button_layout)

    def reset_to_defaults(self):
        """Reset all values to application defaults"""
        self.mz_tolerance_spin.setValue(5.0)
        self.separate_groups_cb.setChecked(True)
        self.rt_shift_spin.setValue(1.0)
        self.crop_rt_cb.setChecked(False)
        self.normalize_cb.setChecked(False)

    def get_values(self):
        """Get the current values from the dialog"""
        return {
            "mz_tolerance_ppm": self.mz_tolerance_spin.value(),
            "separate_groups": self.separate_groups_cb.isChecked(),
            "rt_shift_min": self.rt_shift_spin.value(),
            "crop_rt_window": self.crop_rt_cb.isChecked(),
            "normalize_samples": self.normalize_cb.isChecked(),
        }


def main():
    app = QApplication(sys.argv)

    # Set application properties
    app.setApplicationName("mzML Explorer")
    app.setApplicationVersion("1.0.0")

    # Create and show main window
    window = MzMLExplorerMainWindow()
    window.show()

    sys.exit(app.exec())


if __name__ == "__main__":
    main()
